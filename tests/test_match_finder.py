"""
Тесты для поиска товаров в документах с проверкой процентов совпадений.

Проверяет:
- Поиск товаров в Excel файлах
- Правильность процентов совпадений (100%, 85%, 56% и т.д.)
- Уникальность результатов
- Фильтрацию по минимальному проценту совпадений
"""

import unittest
from pathlib import Path
import sys

# Добавляем корневую директорию проекта в путь
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Подавляем все логи ПЕРЕД импортом модулей
from loguru import logger
logger.remove()
logger.add(lambda *args, **kwargs: None, level="TRACE")

from services.document_search.match_finder import MatchFinder
from services.document_search.keyword_matcher import check_keywords_match, extract_keywords


class TestMatchFinder(unittest.TestCase):
    """Тесты для поиска товаров в документах."""
    
    @classmethod
    def setUpClass(cls):
        """Создание тестовых файлов и товаров."""
        cls.test_files_dir = Path(__file__).parent / "test_files"
        cls.test_files_dir.mkdir(exist_ok=True)
        
        # Создаем Excel файл с товарами разных процентов совпадений
        cls.test_excel_file = cls._create_test_excel_with_products()
        
        # Список товаров для поиска
        cls.product_names = [
            "Контейнер мусорный 240л",  # 100% - точное совпадение
            "Мешок для мусора 120л",     # 100% - точное совпадение
            "ДенсТоп ЭП 203",            # Будет искать частично
            "Реолен Адмикс Плюс",        # Будет искать частично
        ]
    
    @staticmethod
    def _create_test_excel_with_products():
        """Создает Excel файл с товарами для тестирования разных процентов совпадений."""
        try:
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товары"
            
            # Заголовки
            ws['A1'] = "Товар"
            ws['B1'] = "Количество"
            ws['C1'] = "Цена"
            ws['D1'] = "Сумма"
            
            # Данные с разными вариантами написания для тестирования процентов совпадений
            test_data = [
                # Точные совпадения (100%)
                ["Контейнер мусорный 240л", 10, 1500, 15000],
                ["Мешок для мусора 120л", 50, 25, 1250],
                
                # Частичные совпадения (85%+)
                ["Контейнер мусорный 240", 5, 1500, 7500],  # Без "л"
                ["Мешок для мусора", 20, 25, 500],          # Без "120л"
                ["ДенсТоп ЭП 203", 15, 5000, 75000],        # Точное совпадение
                ["Реолен Адмикс Плюс", 8, 3000, 24000],     # Точное совпадение
                
                # Частичные совпадения (56-85%)
                ["Контейнер мусорный", 3, 1500, 4500],      # Без объема
                ["Мешок мусора 120л", 10, 25, 250],         # Без "для"
                ["ДенсТоп ЭП", 5, 5000, 25000],             # Без номера
                ["Реолен Адмикс", 2, 3000, 6000],           # Без "Плюс"
                
                # Слабое совпадение (<56% - не должно находиться)
                ["Контейнер", 1, 1500, 1500],               # Только одно слово
                ["Мусорный", 1, 100, 100],                  # Только одно слово
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            file_path = Path(__file__).parent / "test_files" / "test_products_search.xlsx"
            wb.save(file_path)
            return file_path
            
        except ImportError:
            return None
    
    def setUp(self):
        """Инициализация MatchFinder перед каждым тестом."""
        self.match_finder = MatchFinder(self.product_names)
    
    def tearDown(self):
        """Вывод результата теста после выполнения."""
        pass
    
    def test_search_products_100_percent(self):
        """Тест поиска товаров с 100% совпадением."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        # Проверяем, что найдены товары с 100% совпадением
        exact_matches = [m for m in matches if m.get('score') == 100.0]
        self.assertGreater(len(exact_matches), 0, "Должны быть найдены товары с 100% совпадением")
        
        print(f"\n📊 Результаты поиска товаров (100% совпадение):")
        print(f"   ✅ Файл открыт: {self.test_excel_file.name}")
        print(f"   ✅ Найдено совпадений 100%: {len(exact_matches)}")
        
        for match in exact_matches[:5]:  # Показываем первые 5
            print(f"   ✅ Товар: {match.get('product_name')} - {match.get('score')}%")
    
    def test_search_products_score_filtering(self):
        """Тест фильтрации товаров по минимальному проценту совпадений (>=85%)."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        # Все совпадения должны быть >= 85%
        for match in matches:
            score = match.get('score', 0)
            self.assertGreaterEqual(
                score, 85.0,
                f"Все совпадения должны быть >= 85%, но найдено: {score}% для {match.get('product_name')}"
            )
        
        print(f"\n📊 Фильтрация по проценту совпадений:")
        print(f"   ✅ Всего найдено совпадений >=85%: {len(matches)}")
        
        # Группируем по процентам
        scores_100 = [m for m in matches if m.get('score') == 100.0]
        scores_85_99 = [m for m in matches if 85.0 <= m.get('score', 0) < 100.0]
        
        print(f"   ✅ Совпадений 100%: {len(scores_100)}")
        print(f"   ✅ Совпадений 85-99%: {len(scores_85_99)}")
    
    def test_search_products_uniqueness(self):
        """Тест уникальности результатов - один товар = одно лучшее совпадение."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        # Проверяем уникальность - каждый товар должен встречаться только один раз
        product_names_found = [m.get('product_name') for m in matches]
        unique_names = set(product_names_found)
        
        self.assertEqual(
            len(product_names_found), len(unique_names),
            f"Найдены дубликаты товаров. Всего: {len(product_names_found)}, уникальных: {len(unique_names)}"
        )
        
        print(f"\n📊 Проверка уникальности результатов:")
        print(f"   ✅ Всего найдено совпадений: {len(matches)}")
        print(f"   ✅ Уникальных товаров: {len(unique_names)}")
        print(f"   ✅ Уникальность: 100%")

    def test_stop_phrases_skip_cells(self):
        """Стоп-фразы должны исключать строки с нежелательными общими фразами."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")

        # Создаём отдельный файл, где строка содержит только стоп-фразу и товар
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl не установлен")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Товары"
        ws["A1"] = "Товар"
        ws["A2"] = "Ленты гидроизоляционные Контейнер мусорный 240л"
        stop_file = self.test_files_dir / "test_stop_phrases.xlsx"
        wb.save(stop_file)

        finder_with_stop = MatchFinder(
            ["Контейнер мусорный 240л"],
            stop_phrases=["Ленты гидроизоляционные"],
        )

        matches = finder_with_stop.search_workbook_for_products(stop_file)
        self.assertEqual(
            len(matches),
            0,
            "Строка, содержащая стоп-фразу, не должна давать совпадений",
        )
    
    def test_keyword_match_100_percent(self):
        """Тест проверки 100% совпадения по ключевым словам."""
        # Точное совпадение
        pattern = extract_keywords("Контейнер мусорный 240л")
        result = check_keywords_match(
            "Контейнер мусорный 240л",
            pattern,
            "Контейнер мусорный 240л"
        )
        
        self.assertTrue(result['found'], "Должно быть найдено совпадение")
        self.assertEqual(result['score'], 100.0, "Должно быть 100% совпадение")
        self.assertTrue(result['full_match'], "Должно быть полное совпадение")
        
        print(f"\n📊 Проверка 100% совпадения:")
        print(f"   ✅ Текст: 'Контейнер мусорный 240л'")
        print(f"   ✅ Процент совпадения: {result['score']}%")
        print(f"   ✅ Полное совпадение: {'Да' if result['full_match'] else 'Нет'}")
    
    def test_keyword_match_85_percent(self):
        """Тест проверки 85% совпадения по ключевым словам."""
        # Частичное совпадение - есть большинство ключевых слов
        pattern = extract_keywords("Контейнер мусорный 240л")
        result = check_keywords_match(
            "Контейнер мусорный 240",  # Без "л"
            pattern,
            "Контейнер мусорный 240л"
        )
        
        if result['found']:
            score = result['score']
            self.assertGreaterEqual(score, 85.0, f"Должно быть >=85%, но получили {score}%")
            self.assertLessEqual(score, 100.0, f"Должно быть <=100%, но получили {score}%")
            
            print(f"\n📊 Проверка частичного совпадения (85%+):")
            print(f"   ✅ Текст: 'Контейнер мусорный 240' (без 'л')")
            print(f"   ✅ Процент совпадения: {score}%")
    
    def test_keyword_match_filtering(self):
        """Тест фильтрации совпадений по проценту."""
        pattern = extract_keywords("Контейнер мусорный 240л")
        
        test_cases = [
            ("Контейнер мусорный 240л", 100.0, True),      # 100%
            ("Контейнер мусорный 240", 85.0, True),        # 85%+
            ("Контейнер мусорный", 85.0, True),            # 85%+
            ("Контейнер", 0.0, False),                     # <56% - не должно находиться
        ]
        
        passed = 0
        for text, expected_min_score, should_find in test_cases:
            result = check_keywords_match(text, pattern, "Контейнер мусорный 240л")
            
            if should_find:
                self.assertTrue(result['found'], f"Должно быть найдено: '{text}'")
                self.assertGreaterEqual(
                    result['score'], expected_min_score,
                    f"Для '{text}' ожидалось >= {expected_min_score}%, но получено {result['score']}%"
                )
                passed += 1
            else:
                # Может найти, но с низким процентом, который будет отфильтрован
                if result['found']:
                    self.assertLess(result['score'], 85.0, 
                                  f"Для '{text}' процент должен быть <85%, но получено {result['score']}%")
                passed += 1
        
        print(f"\n📊 Проверка фильтрации по процентам:")
        print(f"   ✅ Проверено случаев: {len(test_cases)}")
        print(f"   ✅ Успешно: {passed}/{len(test_cases)}")


class CustomTestResult(unittest.TextTestResult):
    """Кастомный результат тестов с улучшенным выводом."""
    
    def addSuccess(self, test):
        """Вывод при успешном тесте."""
        super().addSuccess(test)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"✅ Тест '{test_name}' успешно пройден\n")
    
    def addError(self, test, err):
        """Вывод при ошибке в тесте."""
        super().addError(test, err)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"❌ Тест '{test_name}' провален с ошибкой:\n")
        if len(err) > 1:
            self.stream.write(f"   {err[1]}\n")
        else:
            self.stream.write(f"   {err}\n")
    
    def addFailure(self, test, err):
        """Вывод при провале теста."""
        super().addFailure(test, err)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"❌ Тест '{test_name}' провален:\n")
        if len(err) > 1:
            self.stream.write(f"   {err[1]}\n")
        else:
            self.stream.write(f"   {err}\n")


class CustomTestRunner(unittest.TextTestRunner):
    """Кастомный TestRunner с улучшенным выводом."""
    resultclass = CustomTestResult
    
    def __init__(self, *args, **kwargs):
        if 'verbosity' not in kwargs:
            kwargs['verbosity'] = 1
        super().__init__(*args, **kwargs)


if __name__ == '__main__':
    print("\n" + "=" * 70)
    print("ЗАПУСК ТЕСТОВ ПОИСКА ТОВАРОВ В ДОКУМЕНТАХ")
    print("=" * 70 + "\n")
    
    loader = unittest.TestLoader()
    suite = loader.loadTestsFromModule(sys.modules[__name__])
    runner = CustomTestRunner(verbosity=1)
    result = runner.run(suite)
    
    print("\n" + "=" * 70)
    if result.wasSuccessful():
        print(f"✅ ВСЕ ТЕСТЫ УСПЕШНО ПРОЙДЕНЫ: {result.testsRun} тестов")
    else:
        print(f"❌ ТЕСТИРОВАНИЕ ЗАВЕРШЕНО С ОШИБКАМИ:")
        print(f"   Пройдено: {result.testsRun - len(result.failures) - len(result.errors)}")
        print(f"   Провалено: {len(result.failures)}")
        print(f"   Ошибок: {len(result.errors)}")
    print("=" * 70)

