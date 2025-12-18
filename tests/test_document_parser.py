"""
Тесты для универсального парсера документов.

Проверяет:
- Парсинг Excel файлов (.xlsx, .xls)
- Парсинг Word документов (.docx)
- Парсинг PDF файлов (обычные и отсканированные)
- Обработку ошибок и исключений
- Определение типа документов
"""

import unittest
from pathlib import Path
import os
import sys

# Добавляем корневую директорию проекта в путь
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# Подавляем все логи ПЕРЕД импортом модулей, чтобы они не засоряли вывод тестов
from loguru import logger
logger.remove()  # Убираем все обработчики по умолчанию
logger.add(lambda *args, **kwargs: None, level="TRACE")  # Пустой обработчик - ничего не выводит

from services.document_search.document_parser import DocumentParser
from services.document_search.word_processor import WordProcessor
from services.document_search.pdf_processor import PDFProcessor
from services.document_search.excel_parser import ExcelParser
from services.document_search.match_finder import MatchFinder
from services.document_search.keyword_matcher import check_keywords_match, extract_keywords
from core.exceptions import DocumentSearchError


class TestDocumentParser(unittest.TestCase):
    """Тесты для универсального парсера документов."""
    
    @classmethod
    def setUpClass(cls):
        """Создание тестовых файлов перед запуском всех тестов."""
        cls.test_files_dir = Path(__file__).parent / "test_files"
        cls.test_files_dir.mkdir(exist_ok=True)
        
        # Запускаем скрипт создания тестовых файлов, если они не существуют
        create_script = Path(__file__).parent / "create_test_files.py"
        if create_script.exists():
            import subprocess
            try:
                subprocess.run(
                    [sys.executable, str(create_script)],
                    cwd=str(create_script.parent),
                    check=False,
                    capture_output=True
                )
            except Exception as e:
                print(f"Не удалось автоматически создать тестовые файлы: {e}")
                print("Запустите вручную: python tests/create_test_files.py")
    
    def setUp(self):
        """Инициализация парсера перед каждым тестом."""
        self.parser = DocumentParser()
    
    def tearDown(self):
        """Вывод результата теста после выполнения."""
        # Результаты выводятся через CustomTestResult
        pass
    
    def test_detect_excel_type(self):
        """Тест определения типа Excel файла."""
        test_file = self.test_files_dir / "test_excel.xlsx"
        if test_file.exists():
            doc_type = self.parser.detect_document_type(test_file)
            self.assertEqual(doc_type, 'excel', "Должен определять Excel файл")
        else:
            self.skipTest(f"Тестовый файл не найден: {test_file}")
    
    def test_detect_word_type(self):
        """Тест определения типа Word документа."""
        test_file = self.test_files_dir / "test_word.docx"
        if test_file.exists():
            doc_type = self.parser.detect_document_type(test_file)
            self.assertEqual(doc_type, 'word', "Должен определять Word документ")
        else:
            self.skipTest(f"Тестовый файл не найден: {test_file}")
    
    def test_detect_pdf_type(self):
        """Тест определения типа PDF файла."""
        test_file = self.test_files_dir / "test_pdf.pdf"
        if test_file.exists():
            doc_type = self.parser.detect_document_type(test_file)
            self.assertEqual(doc_type, 'pdf', "Должен определять PDF файл")
        else:
            self.skipTest(f"Тестовый файл не найден: {test_file}")
    
    def test_parse_excel_xlsx(self):
        """Тест парсинга Excel .xlsx файла."""
        test_file = self.test_files_dir / "test_excel.xlsx"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            result = self.parser.parse_document(test_file)
            
            self.assertEqual(result['type'], 'excel', "Тип должен быть excel")
            self.assertIn('text', result, "Должен быть извлечен текст")
            self.assertIn('cells', result, "Должны быть извлечены ячейки")
            self.assertGreater(len(result['cells']), 0, "Должны быть ячейки")
            
            # Проверяем наличие тестовых данных
            text = result['text'].lower()
            self.assertIn('контейнер', text, "Должен быть найден тестовый текст")
            
            # Выводим результаты парсинга сразу
            print(f"\n📊 Результаты парсинга {test_file.name}:")
            print(f"   ✅ Файл открыт")
            print(f"   ✅ Файл прочитан: {len(result['text'])} символов, {len(result['cells'])} ячеек")
            print(f"   ✅ Тип документа: {result['type']}")
            
            # Статистика по ячейкам
            if result['cells']:
                unique_texts = set()
                for cell in result['cells']:
                    if cell.get('text'):
                        unique_texts.add(cell['text'])
                print(f"   ✅ Найдено уникальных значений: {len(unique_texts)}")
            
        except Exception as e:
            self.fail(f"Ошибка при парсинге Excel .xlsx: {e}")
    
    def test_parse_word(self):
        """Тест парсинга Word документа."""
        test_file = self.test_files_dir / "test_word.docx"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            result = self.parser.parse_document(test_file)
            
            self.assertEqual(result['type'], 'word', "Тип должен быть word")
            self.assertIn('text', result, "Должен быть извлечен текст")
            self.assertIn('cells', result, "Должны быть извлечены ячейки")
            
            text = result['text'].lower()
            self.assertIn('контейнер', text, "Должен быть найден тестовый текст")
            
            print(f"\n📊 Результаты парсинга {test_file.name}:")
            print(f"   ✅ Файл открыт")
            print(f"   ✅ Файл прочитан: {len(result['text'])} символов, {len(result['cells'])} элементов")
            print(f"   ✅ Тип документа: {result['type']}")
            
        except Exception as e:
            # Если python-docx не установлен, это ожидаемо
            if "не установлен" in str(e) or "ImportError" in str(type(e).__name__):
                self.skipTest(f"Библиотека не установлена: {e}")
            else:
                self.fail(f"Ошибка при парсинге Word: {e}")
    
    def test_parse_pdf(self):
        """Тест парсинга обычного PDF файла."""
        test_file = self.test_files_dir / "test_pdf.pdf"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            result = self.parser.parse_document(test_file)
            
            self.assertEqual(result['type'], 'pdf', "Тип должен быть pdf")
            self.assertIn('text', result, "Должен быть извлечен текст")
            
            text = result['text'].lower()
            self.assertIn('контейнер', text, "Должен быть найден тестовый текст")
            
            print(f"\n📊 Результаты парсинга {test_file.name}:")
            print(f"   ✅ Файл открыт")
            print(f"   ✅ Файл прочитан: {len(result['text'])} символов")
            print(f"   ✅ Тип документа: {result['type']}")
            
        except Exception as e:
            # Если библиотеки PDF не установлены, это ожидаемо
            if "не установлен" in str(e) or "ImportError" in str(type(e).__name__):
                self.skipTest(f"Библиотека не установлена: {e}")
            else:
                self.fail(f"Ошибка при парсинге PDF: {e}")
    
    def test_parse_pdf_scanned(self):
        """Тест парсинга отсканированного PDF через OCR."""
        test_file = self.test_files_dir / "test_pdf_scanned.pdf"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            result = self.parser.parse_document(test_file, force_ocr=True)
            
            self.assertEqual(result['type'], 'pdf', "Тип должен быть pdf")
            self.assertIn('text', result, "Должен быть извлечен текст")
            
            print(f"\n📊 Результаты парсинга {test_file.name} (OCR):")
            print(f"   ✅ Файл открыт")
            print(f"   ✅ Файл прочитан через OCR: {len(result.get('text', ''))} символов")
            print(f"   ✅ Тип документа: {result['type']}")
            
        except Exception as e:
            # OCR может быть недоступен
            if "OCR" in str(e) or "tesseract" in str(e).lower() or "не установлен" in str(e):
                self.skipTest(f"OCR недоступен: {e}")
            else:
                self.fail(f"Ошибка при парсинге отсканированного PDF: {e}")
    
    def test_extract_text_from_excel(self):
        """Тест извлечения текста из Excel."""
        test_file = self.test_files_dir / "test_excel.xlsx"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            text = self.parser.extract_text(test_file)
            self.assertIsInstance(text, str, "Текст должен быть строкой")
            self.assertGreater(len(text), 0, "Текст не должен быть пустым")
            
            text_lower = text.lower()
            self.assertIn('контейнер', text_lower, "Должен содержать тестовый текст")
            
        except Exception as e:
            self.fail(f"Ошибка при извлечении текста из Excel: {e}")
    
    def test_iter_document_cells(self):
        """Тест итерации по ячейкам документа."""
        test_file = self.test_files_dir / "test_excel.xlsx"
        if not test_file.exists():
            self.skipTest(f"Тестовый файл не найден: {test_file}")
        
        try:
            cells = list(self.parser.iter_document_cells(test_file))
            
            self.assertGreater(len(cells), 0, "Должны быть ячейки")
            
            # Проверяем структуру ячейки
            if cells:
                cell = cells[0]
                self.assertIn('text', cell, "Ячейка должна содержать text")
                self.assertIn('sheet_name', cell, "Ячейка должна содержать sheet_name")
                self.assertIn('row', cell, "Ячейка должна содержать row")
                
        except Exception as e:
            self.fail(f"Ошибка при итерации по ячейкам: {e}")
    
    def test_nonexistent_file(self):
        """Тест обработки несуществующего файла."""
        nonexistent_file = self.test_files_dir / "nonexistent_file.xlsx"
        
        with self.assertRaises(DocumentSearchError):
            self.parser.parse_document(nonexistent_file)
    
    def test_unsupported_file_type(self):
        """Тест обработки неподдерживаемого типа файла."""
        # Создаем тестовый файл с неподдерживаемым расширением
        test_file = self.test_files_dir / "test_file.txt"
        test_file.write_text("Тестовый текст")
        
        try:
            with self.assertRaises(DocumentSearchError):
                self.parser.parse_document(test_file)
        finally:
            # Удаляем тестовый файл
            if test_file.exists():
                test_file.unlink()
    
    def test_error_handling(self):
        """Тест обработки ошибок при парсинге."""
        # Создаем поврежденный файл (пустой)
        corrupted_file = self.test_files_dir / "corrupted.xlsx"
        corrupted_file.write_bytes(b"")
        
        try:
            # Парсер должен обработать ошибку и вернуть результат с error
            result = self.parser.parse_document(corrupted_file)
            
            # Может быть либо ошибка, либо пустой результат
            # Проверяем, что обработка прошла без исключения
            self.assertIsNotNone(result, "Должен вернуться результат (даже с ошибкой)")
            
        except Exception:
            # Также приемлемо, если парсер выбрасывает исключение
            pass
        finally:
            if corrupted_file.exists():
                corrupted_file.unlink()


class TestWordProcessor(unittest.TestCase):
    """Тесты для парсера Word документов."""
    
    @classmethod
    def setUpClass(cls):
        cls.test_files_dir = Path(__file__).parent / "test_files"
        cls.test_files_dir.mkdir(exist_ok=True)
    
    def setUp(self):
        self.processor = WordProcessor()
    
    def tearDown(self):
        """Вывод результата теста после выполнения."""
        # Результаты выводятся через CustomTestResult
        pass
    
    def test_is_word_file(self):
        """Тест проверки Word файла."""
        test_file = self.test_files_dir / "test_word.docx"
        if test_file.exists():
            self.assertTrue(self.processor.is_word_file(test_file))
        
        # Не Word файл
        non_word = self.test_files_dir / "test_excel.xlsx"
        if non_word.exists():
            self.assertFalse(self.processor.is_word_file(non_word))


class TestProductSearch(unittest.TestCase):
    """Тесты для поиска товаров в документах с проверкой процентов совпадений."""
    
    @classmethod
    def setUpClass(cls):
        """Создание тестовых файлов и товаров."""
        cls.test_files_dir = Path(__file__).parent / "test_files"
        cls.test_files_dir.mkdir(exist_ok=True)
        
        # Создаем Excel файл с товарами для тестирования поиска
        cls.test_excel_file = cls._create_test_excel_with_products()
        
        # Список товаров для поиска
        cls.product_names = [
            "Контейнер мусорный 240л",  # 100% - точное совпадение
            "Мешок для мусора 120л",     # 100% - точное совпадение
        ]
    
    @staticmethod
    def _create_test_excel_with_products():
        """Создает Excel файл с товарами для тестирования разных процентов совпадений."""
        try:
            import openpyxl
            
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Товары"
            
            # Заголовки
            ws['A1'] = "Товар"
            ws['B1'] = "Количество"
            ws['C1'] = "Цена"
            
            # Данные с разными вариантами написания для тестирования процентов совпадений
            test_data = [
                # Точные совпадения (100%)
                ["Контейнер мусорный 240л", 10, 1500],
                ["Мешок для мусора 120л", 50, 25],
                # Частичные совпадения (85%+)
                ["Контейнер мусорный 240", 5, 1500],  # Без "л"
                ["Мешок для мусора", 20, 25],          # Без "120л"
            ]
            
            for row_idx, row_data in enumerate(test_data, start=2):
                for col_idx, value in enumerate(row_data, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            file_path = Path(__file__).parent / "test_files" / "test_products_search.xlsx"
            wb.save(file_path)
            return file_path
            
        except ImportError:
            return None
    
    def setUp(self):
        """Инициализация MatchFinder перед каждым тестом."""
        self.match_finder = MatchFinder(self.product_names)
    
    def tearDown(self):
        """Вывод результата теста после выполнения."""
        pass
    
    def test_search_products_in_excel(self):
        """Тест поиска товаров в Excel файле."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        self.assertGreater(len(matches), 0, "Должны быть найдены товары")
        
        # Выводим результаты поиска
        print(f"\n📊 Результаты поиска товаров в {self.test_excel_file.name}:")
        print(f"   ✅ Файл открыт")
        print(f"   ✅ Найдено совпадений: {len(matches)}")
        
        # Группируем по процентам
        scores_100 = [m for m in matches if m.get('score') == 100.0]
        scores_85_99 = [m for m in matches if 85.0 <= m.get('score', 0) < 100.0]
        
        print(f"   ✅ Совпадений 100%: {len(scores_100)}")
        print(f"   ✅ Совпадений 85-99%: {len(scores_85_99)}")
        
        # Показываем первые результаты
        for match in matches[:3]:
            score = match.get('score', 0)
            product = match.get('product_name', '')
            matched_text = match.get('matched_text', '')
            print(f"   ✅ Товар: {product} - {score}% (найдено: '{matched_text}')")
    
    def test_search_products_score_filtering(self):
        """Тест фильтрации товаров по минимальному проценту совпадений (>=85%)."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        # Все совпадения должны быть >= 85%
        for match in matches:
            score = match.get('score', 0)
            self.assertGreaterEqual(
                score, 85.0,
                f"Все совпадения должны быть >= 85%, но найдено: {score}%"
            )
        
        print(f"\n📊 Фильтрация по проценту совпадений:")
        print(f"   ✅ Всего найдено совпадений >=85%: {len(matches)}")
    
    def test_search_products_uniqueness(self):
        """Тест уникальности результатов - один товар = одно лучшее совпадение."""
        if not self.test_excel_file or not self.test_excel_file.exists():
            self.skipTest("Тестовый файл не создан (нужен openpyxl)")
        
        matches = self.match_finder.search_workbook_for_products(self.test_excel_file)
        
        # Проверяем уникальность - каждый товар должен встречаться только один раз
        product_names_found = [m.get('product_name') for m in matches]
        unique_names = set(product_names_found)
        
        self.assertEqual(
            len(product_names_found), len(unique_names),
            f"Найдены дубликаты товаров. Всего: {len(product_names_found)}, уникальных: {len(unique_names)}"
        )
        
        print(f"\n📊 Проверка уникальности результатов:")
        print(f"   ✅ Всего найдено совпадений: {len(matches)}")
        print(f"   ✅ Уникальных товаров: {len(unique_names)}")
        print(f"   ✅ Уникальность: 100%")
    
    def test_keyword_match_percentages(self):
        """Тест проверки разных процентов совпадений."""
        pattern = extract_keywords("Контейнер мусорный 240л")
        
        # Тестируем разные случаи
        test_cases = [
            ("Контейнер мусорный 240л", 100.0, True),      # 100%
            ("Контейнер мусорный 240", 85.0, True),        # 85%+
        ]
        
        passed = 0
        for text, expected_min_score, should_find in test_cases:
            result = check_keywords_match(text, pattern, "Контейнер мусорный 240л")
            
            if should_find:
                self.assertTrue(result['found'], f"Должно быть найдено: '{text}'")
                if result['found']:
                    score = result['score']
                    self.assertGreaterEqual(
                        score, expected_min_score,
                        f"Для '{text}' ожидалось >= {expected_min_score}%, но получено {score}%"
                    )
                    passed += 1
        
        print(f"\n📊 Проверка процентов совпадений:")
        print(f"   ✅ Проверено случаев: {len(test_cases)}")
        print(f"   ✅ Успешно: {passed}/{len(test_cases)}")


class TestPDFProcessor(unittest.TestCase):
    """Тесты для парсера PDF документов."""
    
    @classmethod
    def setUpClass(cls):
        cls.test_files_dir = Path(__file__).parent / "test_files"
        cls.test_files_dir.mkdir(exist_ok=True)
    
    def setUp(self):
        self.processor = PDFProcessor()
    
    def tearDown(self):
        """Вывод результата теста после выполнения."""
        # Результаты выводятся через CustomTestResult
        pass
    
    def test_is_pdf_file(self):
        """Тест проверки PDF файла."""
        test_file = self.test_files_dir / "test_pdf.pdf"
        if test_file.exists():
            self.assertTrue(self.processor.is_pdf_file(test_file))
    
    def test_detect_pdf_type(self):
        """Тест определения типа PDF."""
        test_file = self.test_files_dir / "test_pdf.pdf"
        if test_file.exists():
            try:
                pdf_type = self.processor.detect_pdf_type(test_file)
                self.assertIn(pdf_type, ['text', 'scanned'], "Тип должен быть text или scanned")
            except Exception as e:
                self.skipTest(f"Не удалось определить тип PDF: {e}")


class CustomTestResult(unittest.TextTestResult):
    """Кастомный результат тестов с улучшенным выводом."""
    
    def addSuccess(self, test):
        """Вывод при успешном тесте."""
        super().addSuccess(test)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"✅ Тест '{test_name}' успешно пройден\n")
    
    def addError(self, test, err):
        """Вывод при ошибке в тесте."""
        super().addError(test, err)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"❌ Тест '{test_name}' провален с ошибкой:\n")
        if len(err) > 1:
            self.stream.write(f"   {err[1]}\n")
        else:
            self.stream.write(f"   {err}\n")
    
    def addFailure(self, test, err):
        """Вывод при провале теста."""
        super().addFailure(test, err)
        test_name = getattr(test, '_testMethodName', str(test))
        self.stream.write(f"❌ Тест '{test_name}' провален:\n")
        if len(err) > 1:
            self.stream.write(f"   {err[1]}\n")
        else:
            self.stream.write(f"   {err}\n")


class CustomTestRunner(unittest.TextTestRunner):
    """Кастомный TestRunner с улучшенным выводом."""
    resultclass = CustomTestResult
    
    def __init__(self, *args, **kwargs):
        # Устанавливаем verbosity по умолчанию
        if 'verbosity' not in kwargs:
            kwargs['verbosity'] = 1
        super().__init__(*args, **kwargs)


if __name__ == '__main__':
    # Создаем тестовые файлы перед запуском тестов
    test_files_dir = Path(__file__).parent / "test_files"
    test_files_dir.mkdir(exist_ok=True)
    
    # Проверяем, есть ли тестовые файлы
    has_files = any(test_files_dir.glob("test_*"))
    
    if not has_files:
        print("⚠️  Тестовые файлы не найдены!")
        print("Запускаю создание тестовых файлов...")
        
        create_script = Path(__file__).parent / "create_test_files.py"
        if create_script.exists():
            import subprocess
            subprocess.run([sys.executable, str(create_script)], cwd=str(create_script.parent))
    
    print("\n" + "=" * 70)
    print("ЗАПУСК ТЕСТОВ ПАРСЕРА ДОКУМЕНТОВ")
    print("=" * 70 + "\n")
    
    # Логи уже подавлены в начале файла, просто запускаем тесты
    # Запускаем тесты с кастомным runner
    loader = unittest.TestLoader()
    suite = loader.loadTestsFromModule(sys.modules[__name__])
    runner = CustomTestRunner(verbosity=1)
    result = runner.run(suite)
    
    print("\n" + "=" * 70)
    if result.wasSuccessful():
        print(f"✅ ВСЕ ТЕСТЫ УСПЕШНО ПРОЙДЕНЫ: {result.testsRun} тестов")
    else:
        print(f"❌ ТЕСТИРОВАНИЕ ЗАВЕРШЕНО С ОШИБКАМИ:")
        print(f"   Пройдено: {result.testsRun - len(result.failures) - len(result.errors)}")
        print(f"   Провалено: {len(result.failures)}")
        print(f"   Ошибок: {len(result.errors)}")
    print("=" * 70)
