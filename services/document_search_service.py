"""
Сервис для скачивания документации торгов и поиска товаров внутри XLSX-файлов.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import uuid
import os

import requests
from loguru import logger
from openpyxl import load_workbook
from rapidfuzz import fuzz, process
import rarfile
import zipfile
import py7zr

from core.database import DatabaseManager
from core.exceptions import DocumentSearchError


class DocumentSearchService:
    """
    Сервис поиска информации в документации торгов.

    Выполняет:
    - Поиск подходящего документа (содержит слово «смета»)
    - Скачивание файла в указанную директорию
    - Парсинг XLSX и нечеткий поиск упоминаний товаров
    """

    KEYWORD_PATTERN = re.compile(r"смет\w*", re.IGNORECASE)
    ARCHIVE_PATTERN = re.compile(
        r"^(?P<base>.+?)(?:[._ -]*(?:part)?(?P<part>\d+))?\.(rar|zip|7z)$",
        re.IGNORECASE,
    )

    DEFAULT_HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://zakupki.gov.ru/",
        "Connection": "keep-alive",
    }

    def __init__(
        self,
        db_manager: DatabaseManager,
        download_dir: Path,
        unrar_path: Optional[str] = None,
    ):
        """
        Args:
            db_manager: Менеджер БД с таблицей products
            download_dir: Директория для сохранения файлов
            unrar_path: Путь к инструменту UnRAR (опционально)
        """
        self.db_manager = db_manager
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self._product_names: Optional[List[str]] = None
        self._keyword = "смета"
        self.http_session = requests.Session()
        self.http_session.headers.update(self.DEFAULT_HEADERS)
        self._rar_tool_configured = False
        self._unrar_path = Path(unrar_path) if unrar_path else None
        
        # Установка пути к WinRAR в переменную среды PATH
        # Путь к WinRAR настраивается через переменную окружения WINRAR_PATH в .env файле
        # Пример: WINRAR_PATH=C:\Program Files\WinRAR
        winrar_path = os.environ.get("WINRAR_PATH", r"C:\Program Files\WinRAR")
        if os.path.exists(winrar_path):
            os.environ["PATH"] = os.environ.get("PATH", "") + os.pathsep + winrar_path
        
        logger.debug(f"DocumentSearchService инициализирован, каталог загрузки: {self.download_dir}")

    def ensure_products_loaded(self) -> None:
        """Ленивая загрузка названий товаров (по требованию пользователя)."""
        if self._product_names is not None:
            return

        logger.info("Загрузка списка товаров для поиска по документации...")
        query = "SELECT name FROM products WHERE name IS NOT NULL"
        results = self.db_manager.execute_query(query)
        self._product_names = [row.get("name", "").strip() for row in results if row.get("name")]
        logger.info(f"Получено наименований товаров: {len(self._product_names)}")

    def run_document_search(self, documents: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Основной сценарий: найти документ, скачать и выполнить поиск.

        Args:
            documents: Метаданные документов торга

        Returns:
            Словарь с путем к файлу и найденными совпадениями
        """
        if not documents:
            raise DocumentSearchError("У выбранного торга нет приложенных документов.")

        target_doc = self._choose_document(documents)
        downloaded_paths = self._download_required_documents(target_doc, documents)
        workbook_path = self._prepare_workbook_path(downloaded_paths)
        matches = self._search_workbook_for_products(workbook_path)

        print(f"Документ скачан: {workbook_path}")
        print(f"Результаты парсинга: {matches}")
        logger.info(f"Поиск по документации завершен, найдено совпадений: {len(matches)}")

        return {
            "file_path": str(workbook_path),
            "matches": matches,
        }

    def _choose_document(self, documents: List[Dict[str, Any]]) -> Dict[str, Any]:
        """Выбор документа со словом «смета» в названии/ссылке."""
        matches: List[Dict[str, Any]] = []
        for doc in documents:
            file_name = (doc.get("file_name") or "")
            link = (doc.get("document_links") or "")
            combined = f"{file_name} {link}"
            logger.debug(f"Проверка документа: file_name='{file_name}', link='{link}'")
            if self.KEYWORD_PATTERN.search(combined):
                matches.append(doc)

        if not matches:
            raise DocumentSearchError("Не найден документ со словом «смета».")

        matches.sort(key=self._document_priority)
        selected = matches[0]
        logger.info(f"Выбран документ для анализа: {selected.get('file_name') or selected.get('document_links')}")
        return selected

    def _document_priority(self, doc: Dict[str, Any]) -> Tuple[int, str]:
        """Определение приоритета документов (сначала XLSX, затем архивы)."""
        name = (doc.get("file_name") or "").lower()
        if name.endswith(".xlsx"):
            return (0, name)
        if name.endswith(".xls"):
            return (1, name)
        if name.endswith((".rar", ".zip", ".7z")):
            return (2, name)
        return (3, name)

    def _download_required_documents(
        self,
        primary_doc: Dict[str, Any],
        all_documents: List[Dict[str, Any]],
    ) -> List[Path]:
        """Скачивание основного файла и всех частей архива (если необходимо)."""
        if not self._is_rar_document(primary_doc):
            return [self._download_document(primary_doc)]

        related_docs = self._collect_related_archives(primary_doc, all_documents)
        logger.info(
            f"Для архива найдено частей: {len(related_docs)} "
            f"({', '.join(doc.get('file_name') or '' for doc in related_docs)})",
        )
        paths: List[Path] = []
        batch_size = 8
        total_docs = len(related_docs)
        for start in range(0, total_docs, batch_size):
            end = min(start + batch_size, total_docs)
            chunk = related_docs[start:end]
            logger.info(
                f"Скачивание документов {start + 1}-{end} из {total_docs} параллельно",
            )
            chunk_paths = self._download_documents_batch(chunk)
            paths.extend(chunk_paths)
        return paths

    def _download_documents_batch(self, documents: List[Dict[str, Any]]) -> List[Path]:
        """Параллельная загрузка группы документов с сохранением порядка."""
        ordered_paths: List[Optional[Path]] = [None] * len(documents)
        with ThreadPoolExecutor(max_workers=min(8, len(documents))) as executor:
            future_map = {
                executor.submit(self._download_document, doc): index
                for index, doc in enumerate(documents)
            }

            try:
                for future in as_completed(future_map):
                    index = future_map[future]
                    ordered_paths[index] = future.result()
            except Exception as error:
                failed_future = future
                for fut in future_map:
                    fut.cancel()
                failed_index = future_map[failed_future]
                failed_doc = documents[failed_index]
                file_name = failed_doc.get("file_name") or failed_doc.get("document_links")
                raise DocumentSearchError(f"Не удалось скачать документ {file_name}") from error

        return [path for path in ordered_paths if path]

    def _collect_related_archives(
        self,
        primary_doc: Dict[str, Any],
        all_documents: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Сбор всех частей многофайлового архива."""
        primary_base, primary_part = self._split_archive_name(primary_doc.get("file_name"))
        if not primary_base:
            return [primary_doc]

        related: List[Tuple[int, Dict[str, Any]]] = []
        for doc in all_documents:
            if not self._is_rar_document(doc):
                continue
            base, part = self._split_archive_name(doc.get("file_name"))
            if base == primary_base:
                part_index = part if part is not None else 0
                related.append((part_index, doc))

        if not related:
            return [primary_doc]

        related.sort(key=lambda item: item[0])
        ordered_docs: List[Dict[str, Any]] = []
        seen_names = set()
        for _, doc in related:
            name = doc.get("file_name")
            if name in seen_names:
                continue
            seen_names.add(name)
            ordered_docs.append(doc)

        return ordered_docs

    def _is_rar_document(self, document: Dict[str, Any]) -> bool:
        """Проверка, является ли документ архивом (RAR, ZIP, 7Z)."""
        name = (document.get("file_name") or "").lower()
        return name.endswith((".rar", ".zip", ".7z"))

    def _split_archive_name(self, file_name: Optional[str]) -> Tuple[Optional[str], Optional[int]]:
        """Выделение базовой части имени архива и номера части."""
        if not file_name:
            return None, None
        match = self.ARCHIVE_PATTERN.match(file_name)
        if not match:
            return None, None
        base = match.group("base").casefold()
        part = match.group("part")
        return base, int(part) if part is not None else None

    def _prepare_workbook_path(self, downloaded_paths: List[Path]) -> Path:
        """Определение пути к XLSX файлу (непосредственно или после распаковки архива)."""
        if not downloaded_paths:
            raise DocumentSearchError("Не удалось скачать документ.")

        first_path = downloaded_paths[0]
        suffix = first_path.suffix.lower()

        if suffix in {".rar", ".zip", ".7z"}:
            # Для многофайловых архивов сначала склеиваем части, затем распаковываем
            archive_path = self._combine_multi_part_archive(downloaded_paths) if len(downloaded_paths) > 1 else first_path
            return self._extract_archive(archive_path)

        if suffix in {".xlsx", ".xls"}:
            return first_path

        raise DocumentSearchError("Поддерживаются только XLSX файлы или архивы (RAR, ZIP, 7Z) с XLSX.")

    def _combine_multi_part_archive(self, archive_paths: List[Path]) -> Path:
        """
        Склейка многофайлового архива в один файл.
        
        Для архивов, разбитых на части (part1.rar, part2.rar и т.д.),
        склеивает все части в один файл перед распаковкой.
        
        Args:
            archive_paths: Список путей к частям архива в правильном порядке
            
        Returns:
            Путь к склеенному архиву
        """
        if len(archive_paths) == 1:
            return archive_paths[0]
        
        logger.info(f"Обнаружен многофайловый архив из {len(archive_paths)} частей. Начинаю склейку...")
        
        # Определяем базовое имя и расширение
        first_path = archive_paths[0]
        base_name, _ = self._split_archive_name(first_path.name)
        
        if not base_name:
            # Если не удалось определить базовое имя, используем имя первого файла без номера части
            base_name = first_path.stem.rsplit('.part', 1)[0] if '.part' in first_path.stem.lower() else first_path.stem
        
        # Создаем имя для склеенного архива
        suffix = first_path.suffix.lower()
        combined_path = first_path.parent / f"{base_name}_combined{suffix}"
        
        # Склеиваем все части в один файл
        try:
            with open(combined_path, 'wb') as combined_file:
                for part_path in sorted(archive_paths, key=lambda p: self._get_part_number(p)):
                    logger.debug(f"Добавляю часть: {part_path.name}")
                    with open(part_path, 'rb') as part_file:
                        # Копируем данные по частям для экономии памяти
                        while True:
                            chunk = part_file.read(8192)
                            if not chunk:
                                break
                            combined_file.write(chunk)
            
            logger.info(f"Многофайловый архив успешно склеен: {combined_path.name}")
            return combined_path
            
        except Exception as error:
            logger.error(f"Ошибка при склейке архива: {error}")
            # Если склейка не удалась, пробуем использовать первую часть
            logger.warning(f"Использую первую часть архива: {first_path.name}")
            return first_path
    
    def _get_part_number(self, path: Path) -> int:
        """Извлечение номера части из имени файла."""
        match = self.ARCHIVE_PATTERN.match(path.name)
        if match and match.group("part"):
            return int(match.group("part"))
        # Если номер части не найден, проверяем другие форматы
        name_lower = path.name.lower()
        if '.part' in name_lower:
            # Пытаемся извлечь номер из формата .part01, .part1 и т.д.
            part_match = re.search(r'\.part(\d+)', name_lower)
            if part_match:
                return int(part_match.group(1))
        return 0  # Если номер не найден, считаем первой частью

    def _extract_archive(self, archive_path: Path) -> Path:
        """Распаковка архива (ZIP, RAR, 7Z) и поиск XLSX внутри."""
        logger.info(f"Распаковка архива {archive_path.name}")
        extract_dir = archive_path.parent / f"extract_{archive_path.stem}_{uuid.uuid4().hex[:6]}"
        extract_dir.mkdir(parents=True, exist_ok=True)
        suffix = archive_path.suffix.lower()

        try:
            if suffix == ".zip":
                self._extract_zip_archive(archive_path, extract_dir)
            elif suffix == ".rar":
                self._extract_rar_archive(archive_path, extract_dir)
            elif suffix == ".7z":
                self._extract_7z_archive(archive_path, extract_dir)
            else:
                raise DocumentSearchError(f"Неподдерживаемый формат архива: {suffix}")
        except Exception as error:
            logger.error(f"Ошибка распаковки архива {archive_path.name}: {error}")
            raise DocumentSearchError(f"Ошибка распаковки архива: {error}") from error

        xlsx_file = next(extract_dir.rglob("*.xlsx"), None)
        if not xlsx_file:
            raise DocumentSearchError("В архиве не найден XLSX файл.")

        logger.info(f"Найден XLSX после распаковки: {xlsx_file}")
        return xlsx_file

    def _extract_zip_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка ZIP-архива с правильной обработкой кодировок."""
        logger.debug(f"Распаковка ZIP-архива: {archive_path.name}")
        try:
            with zipfile.ZipFile(archive_path, 'r') as zip_ref:
                for zip_info in zip_ref.infolist():
                    # Попытка декодировать имена файлов с использованием cp866
                    try:
                        zip_info.filename = zip_info.filename.encode('cp437').decode('cp866')
                    except (UnicodeEncodeError, UnicodeDecodeError):
                        # Если не удалось декодировать, оставляем оригинальное имя
                        pass
                    zip_ref.extract(zip_info, extract_dir)
            logger.info(f"ZIP-архив {archive_path.name} успешно распакован")
        except zipfile.BadZipFile as error:
            logger.error(f"Поврежденный ZIP-архив: {error}")
            raise DocumentSearchError("ZIP-архив поврежден или имеет неподдерживаемый формат.") from error
        except Exception as error:
            logger.error(f"Ошибка распаковки ZIP-архива: {error}")
            raise DocumentSearchError(f"Ошибка распаковки ZIP-архива: {error}") from error

    def _extract_rar_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка RAR-архива."""
        logger.debug(f"Распаковка RAR-архива: {archive_path.name}")
        self._ensure_unrar_available()
        try:
            with rarfile.RarFile(str(archive_path), 'r') as archive:
                archive.extractall(path=extract_dir)
            logger.info(f"RAR-архив {archive_path.name} успешно распакован")
        except rarfile.NeedFirstVolume as error:
            logger.error(f"Необходимо начинать извлечение архива из первого тома: {error}")
            raise DocumentSearchError(
                "Необходимо начинать извлечение архива из первого тома. "
                "Убедитесь, что все части архива скачаны."
            ) from error
        except rarfile.RarCannotExec as error:
            logger.error("Не найден инструмент для распаковки RAR.")
            raise DocumentSearchError(
                "Не найден инструмент для распаковки RAR. "
                "Установите WinRAR/UnRAR и задайте путь в переменной окружения UNRAR_TOOL.",
            ) from error
        except rarfile.Error as error:
            logger.error(f"Ошибка распаковки RAR-архива: {error}")
            raise DocumentSearchError("RAR-архив поврежден или имеет неподдерживаемый формат.") from error

    def _extract_7z_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка 7Z-архива."""
        logger.debug(f"Распаковка 7Z-архива: {archive_path.name}")
        try:
            with py7zr.SevenZipFile(archive_path, mode='r') as archive:
                archive.extractall(path=extract_dir)
            logger.info(f"7Z-архив {archive_path.name} успешно распакован")
        except Exception as error:
            logger.error(f"Ошибка распаковки 7Z-архива {archive_path.name}: {error}")
            raise DocumentSearchError(f"Ошибка распаковки 7Z-архива: {error}") from error

    def _ensure_unrar_available(self) -> None:
        """Проверка наличия инструмента распаковки RAR."""
        if self._rar_tool_configured and rarfile.UNRAR_TOOL:
            return

        env_tool = os.environ.get("UNRAR_TOOL")
        # Пути к инструментам распаковки RAR настраиваются через переменные окружения
        # UNRAR_TOOL - путь к UnRAR.exe или WinRAR.exe
        # WINRAR_PATH - путь к директории WinRAR (используется для поиска инструментов)
        winrar_path = os.environ.get("WINRAR_PATH", r"C:\Program Files\WinRAR")
        winrar_path_x86 = r"C:\Program Files (x86)\WinRAR"
        candidate_paths = [
            str(self._unrar_path) if self._unrar_path else None,
            env_tool,
            os.path.join(winrar_path, "UnRAR.exe"),
            os.path.join(winrar_path_x86, "UnRAR.exe"),
            os.path.join(winrar_path, "WinRAR.exe"),
            os.path.join(winrar_path_x86, "WinRAR.exe"),
            # Fallback пути (только в комментариях для справки)
            # r"C:\Program Files\WinRAR\UnRAR.exe",
            # r"C:\Program Files (x86)\WinRAR\UnRAR.exe",
            # r"C:\Program Files\WinRAR\WinRAR.exe",
            # r"C:\Program Files (x86)\WinRAR\WinRAR.exe",
        ]

        for candidate in candidate_paths:
            if not candidate:
                continue
            candidate_path = Path(candidate)
            if candidate_path.exists():
                rarfile.UNRAR_TOOL = str(candidate_path)
                self._rar_tool_configured = True
                logger.info(f"Найден инструмент распаковки RAR: {candidate_path}")
                return

        self._rar_tool_configured = False
        logger.warning(
            "Инструмент распаковки RAR не найден. Установите WinRAR/UnRAR "
            "и задайте путь через переменную окружения UNRAR_TOOL.",
        )
        raise DocumentSearchError(
            "Инструмент для распаковки RAR не найден. "
            "Установите WinRAR/UnRAR и задайте UNRAR_TOOL.",
        )

    def _download_document(self, document: Dict[str, Any]) -> Path:
        """Скачивание XLSX-файла по ссылке."""
        url = document.get("document_links")
        if not url:
            raise DocumentSearchError("У выбранного документа отсутствует ссылка для скачивания.")

        raw_name = document.get("file_name") or f"document_{document.get('id', 'unknown')}"
        file_name = self._sanitize_filename(raw_name)
        suffix = Path(file_name).suffix.lower()
        if not suffix:
            file_name = f"{file_name}.xlsx"

        destination = self.download_dir / file_name
        logger.info(f"Начинаю скачивание документа '{file_name}' по ссылке {url}")

        try:
            response = self.http_session.get(url, timeout=60, stream=True, allow_redirects=True)
            response.raise_for_status()
            with destination.open("wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        file.write(chunk)
        except requests.RequestException as error:
            logger.error(f"Ошибка скачивания документа: {error}")
            raise DocumentSearchError("Не удалось скачать документ.") from error

        logger.info(f"Документ сохранен: {destination}")
        return destination

    def _search_workbook_for_products(self, file_path: Path) -> List[Dict[str, Any]]:
        """Парсинг XLSX и поиск совпадений с названиями товаров."""
        self.ensure_products_loaded()
        if not self._product_names:
            logger.warning("Список товаров пуст, поиск не будет выполнен.")
            return []

        best_matches: Dict[str, Dict[str, Any]] = {}

        for text in self._iter_workbook_texts(file_path):
            matches = process.extract(
                text,
                self._product_names,
                scorer=fuzz.WRatio,
                limit=3,
                score_cutoff=80,
            )
            for match in matches:
                product_name, score = match[0], match[1]
                existing = best_matches.get(product_name)
                if existing and existing["score"] >= score:
                    continue
                best_matches[product_name] = {
                    "product_name": product_name,
                    "score": score,
                    "matched_text": text,
                }

        sorted_matches = sorted(best_matches.values(), key=lambda item: item["score"], reverse=True)
        return sorted_matches[:50]

    def _iter_workbook_texts(self, file_path: Path) -> Iterable[str]:
        """Итерация по всем текстовым значениям XLSX."""
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as error:
            logger.error(f"Не удалось открыть документ {file_path}: {error}")
            raise DocumentSearchError("Поврежденный или неподдерживаемый XLSX файл.") from error

        for sheet in workbook.worksheets:
            logger.debug(f"Сканирование листа: {sheet.title}")
            for row in sheet.iter_rows(values_only=True):
                for value in row:
                    if value is None:
                        continue
                    text = self._normalize_cell_value(value)
                    if text:
                        yield text

        workbook.close()

    @staticmethod
    def _normalize_cell_value(value: Any) -> Optional[str]:
        """Очистка значения ячейки перед поиском."""
        if isinstance(value, str):
            text = value.strip()
        elif isinstance(value, (int, float)):
            text = f"{value}".strip()
        else:
            text = str(value).strip()

        cleaned = re.sub(r"\s+", " ", text)
        return cleaned.casefold() if cleaned else None

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        """Удаление запрещенных символов из имени файла."""
        sanitized = re.sub(r"[<>:\"/\\\\|?*]", "_", name)
        return sanitized.strip() or "document.xlsx"

