"""
Модуль для парсинга Excel файлов.

Класс ExcelParser отвечает за:
- Чтение .xlsx и .xls файлов
- Итерацию по ячейкам
- Извлечение данных из строк
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, Optional
import re
import warnings
import gc

from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Подавляем предупреждения openpyxl о header/footer и стилях
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logger.warning("Библиотека xlrd не установлена. Поддержка старых .xls файлов будет недоступна.")

try:
    import pandas as pd
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logger.warning("Библиотека pandas не установлена. Резервный способ открытия файлов будет недоступен.")

from core.exceptions import DocumentSearchError


class ExcelParser:
    """Класс для парсинга Excel файлов."""

    def iter_workbook_cells(self, file_path: Path) -> Iterable[Dict[str, Any]]:
        """
        Итерация по всем ячейкам Excel (XLSX или XLS) с информацией о позиции.
        
        Поддерживает оба формата:
        - .xlsx через openpyxl
        - .xls через xlrd (с автоматическим fallback, если файл содержит данные .xlsx)
        """
        suffix = file_path.suffix.lower()
        
        if suffix == ".xls":
            yield from self._iter_xls_cells(file_path, allow_fallback=True)
            return
        
        try:
            yield from self._iter_xlsx_cells(file_path)
        except DocumentSearchError as error:
            if XLRD_AVAILABLE:
                logger.warning(
                    "openpyxl не смог открыть %s (%s). Пробую обработать через xlrd.",
                    file_path,
                    error,
                )
                yield from self._iter_xls_cells(file_path, allow_fallback=False)
            else:
                raise

    def extract_row_data(self, file_path: Path, sheet_name: str, row_number: int) -> Dict[str, Any]:
        """
        Извлекает данные из строки Excel, включая столбец "Количество" и другие столбцы.
        
        Поддерживает оба формата: .xlsx и .xls
        """
        suffix = file_path.suffix.lower()
        
        try:
            if suffix == ".xls":
                return self._extract_xls_row_data(file_path, sheet_name, row_number)
            else:
                return self._extract_xlsx_row_data(file_path, sheet_name, row_number)
        except Exception as error:
            logger.debug(f"Ошибка при извлечении данных строки {row_number} из {sheet_name}: {error}")
            return {}
    
    def extract_full_row_with_context(
        self, 
        file_path: Path, 
        sheet_name: str, 
        row_number: int, 
        found_column: str,
        context_cols: int = 3
    ) -> Dict[str, Any]:
        """
        Извлекает полную строку с соседними ячейками слева и справа, а также имена столбцов.
        
        Args:
            file_path: Путь к файлу
            sheet_name: Имя листа
            row_number: Номер строки (1-based)
            found_column: Буква столбца, где найдено совпадение (например, "F")
            context_cols: Количество столбцов слева и справа для включения в контекст
        
        Returns:
            Dict с полями:
            - full_row: список всех значений строки с именами столбцов
            - left_context: ячейки слева от найденной
            - right_context: ячейки справа от найденной
            - column_names: имена столбцов (из заголовков)
        """
        suffix = file_path.suffix.lower()
        
        try:
            if suffix == ".xls":
                return self._extract_full_row_xls(file_path, sheet_name, row_number, found_column, context_cols)
            else:
                return self._extract_full_row_xlsx(file_path, sheet_name, row_number, found_column, context_cols)
        except Exception as error:
            logger.debug(f"Ошибка при извлечении полной строки {row_number} из {sheet_name}: {error}")
            return {}

    def _iter_xls_cells(self, file_path: Path, allow_fallback: bool = True) -> Iterable[Dict[str, Any]]:
        """Итерация по .xls файлам с возможностью fallback на openpyxl и pandas."""
        # Попытка 1: xlrd (если доступен)
        if XLRD_AVAILABLE:
            try:
                workbook = xlrd.open_workbook(str(file_path))
                for sheet in workbook.sheets():
                    logger.debug(f"Сканирование листа: {sheet.name}")
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            cell = sheet.cell(row_idx, col_idx)
                            if cell.value is None or cell.value == "":
                                continue
                            
                            text = self._normalize_cell_value(cell.value)
                            if text:
                                col_letter = self._xls_col_to_letter(col_idx)
                                cell_address = f"{sheet.name}!{col_letter}{row_idx + 1}"
                                
                                yield {
                                    "text": text,
                                    "display_text": self._format_cell_display(cell.value),
                                    "sheet_name": sheet.name,
                                    "row": row_idx + 1,
                                    "column": col_letter,
                                    "cell_address": cell_address,
                                }
                return
            except Exception as xlrd_error:
                error_msg = str(xlrd_error).lower()
                if "xlsx file" in error_msg:
                    # Файл имеет расширение .xls, но это XLSX
                    if allow_fallback:
                        logger.warning(
                            "Файл %s имеет расширение .xls, но содержит формат XLSX. "
                            "Повторная попытка через openpyxl.",
                            file_path,
                        )
                        yield from self._iter_xlsx_cells(file_path)
                        return
                # Если xlrd не смог открыть, пробуем другие методы
                logger.warning(f"xlrd не смог открыть {file_path}: {xlrd_error}. Пробую другие методы.")
        
        # Попытка 2: pandas с движком xlrd (для старых .xls)
        if PANDAS_AVAILABLE:
            try:
                logger.info(f"Попытка открыть {file_path} через pandas с движком xlrd")
                yield from self._iter_pandas_cells(file_path, engine='xlrd')
                return
            except Exception as pandas_xlrd_error:
                logger.debug(f"pandas с xlrd не смог открыть {file_path}: {pandas_xlrd_error}")
        
        # Попытка 3: pandas с движком openpyxl (если файл на самом деле .xlsx)
        if PANDAS_AVAILABLE and allow_fallback:
            try:
                logger.info(f"Попытка открыть {file_path} через pandas с движком openpyxl")
                yield from self._iter_pandas_cells(file_path, engine='openpyxl')
                return
            except Exception as pandas_openpyxl_error:
                logger.debug(f"pandas с openpyxl не смог открыть {file_path}: {pandas_openpyxl_error}")
        
        # Попытка 4: openpyxl напрямую (если файл на самом деле .xlsx с неправильным расширением)
        if allow_fallback:
            try:
                logger.info(f"Попытка открыть {file_path} через openpyxl напрямую")
                yield from self._iter_xlsx_cells(file_path)
                return
            except Exception as openpyxl_error:
                logger.debug(f"openpyxl не смог открыть {file_path}: {openpyxl_error}")
        
        # Если все методы не сработали
        raise DocumentSearchError(
            f"Не удалось открыть файл {file_path} ни одним из доступных методов. "
            f"Попробуйте установить xlrd==1.2.0 для поддержки старых .xls файлов: pip install xlrd==1.2.0"
        )

    def _iter_xlsx_cells(self, file_path: Path) -> Iterable[Dict[str, Any]]:
        """Итерация по .xlsx файлам через openpyxl."""
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as error:
            error_msg = str(error).lower()
            # Если openpyxl говорит, что это старый формат .xls, пробуем pandas
            if "old .xls file format" in error_msg or "does not support" in error_msg:
                if PANDAS_AVAILABLE:
                    logger.warning(
                        "openpyxl не может открыть %s (старый формат .xls). Пробую через pandas.",
                        file_path,
                    )
                    yield from self._iter_pandas_cells(file_path, engine='xlrd')
                    return
            logger.error(f"Не удалось открыть документ {file_path}: {error}")
            raise DocumentSearchError(f"Поврежденный или неподдерживаемый XLSX файл: {error}") from error
        
        try:
            for sheet in workbook.worksheets:
                logger.debug(f"Сканирование листа: {sheet.title}")
                for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    for cell in row:
                        if cell.value is None:
                            continue
                        text = self._normalize_cell_value(cell.value)
                        if text:
                            yield {
                                "text": text,
                                "display_text": self._format_cell_display(cell.value),
                                "sheet_name": sheet.title,
                                "row": row_idx,
                                "column": cell.column_letter,
                                "cell_address": f"{sheet.title}!{cell.coordinate}",
                            }
        finally:
            workbook.close()

    def _extract_xls_row_data(
        self,
        file_path: Path,
        sheet_name: str,
        row_number: int,
        allow_fallback: bool = True,
    ) -> Dict[str, Any]:
        """Извлечение данных из .xls файла."""
        if not XLRD_AVAILABLE:
            return {}
        
        try:
            workbook = xlrd.open_workbook(str(file_path))
        except Exception as error:
            if allow_fallback and "xlsx file" in str(error).lower():
                logger.warning(
                    "Файл %s имеет расширение .xls, но содержит формат XLSX (лист %s). "
                    "Повторная попытка через openpyxl.",
                    file_path,
                    sheet_name,
                )
                return self._extract_xlsx_row_data(file_path, sheet_name, row_number)
            logger.error(f"Не удалось открыть документ {file_path}: {error}")
            return {}
        
        sheet = workbook.sheet_by_name(sheet_name)
        
        row_idx = row_number - 1
        if row_idx >= sheet.nrows:
            return {}
        
        row_values = []
        max_col = sheet.ncols
        for col_idx in range(max_col):
            cell = sheet.cell(row_idx, col_idx)
            row_values.append(self._format_cell_display(cell.value) if cell.value else "")
        
        column_headers = {}
        header_row = None
        
        for header_row_idx in range(min(5, sheet.nrows)):
            for col_idx in range(min(max_col, 20)):
                cell = sheet.cell(header_row_idx, col_idx)
                if cell.value:
                    header_value = str(cell.value).strip().lower()
                    if header_value == "количество" or (
                        header_value.startswith("количество") and "количество" not in column_headers
                    ):
                        col_letter = self._xls_col_to_letter(col_idx)
                        column_headers["количество"] = {"col": col_letter, "idx": col_idx}
                        if header_row is None:
                            header_row = header_row_idx
        
        if "количество" not in column_headers and max_col >= 4:
            column_headers["количество"] = {"col": "D", "idx": 3}
        
        cost_unit_col = None
        total_cost_col = None
        
        for header_row_idx in range(min(5, sheet.nrows)):
            for col_idx in range(min(max_col, 20)):
                cell = sheet.cell(header_row_idx, col_idx)
                if cell.value:
                    header_value = str(cell.value).strip().lower()
                    if ("стоимость единицы" in header_value or 
                        (header_value.startswith("стоимость") and "единицы" in header_value)):
                        if cost_unit_col is None:
                            col_letter = self._xls_col_to_letter(col_idx)
                            cost_unit_col = {"col": col_letter, "idx": col_idx, "name": str(cell.value).strip()}
                    if "общая стоимость" in header_value or header_value.startswith("общая стоимость"):
                        if total_cost_col is None:
                            col_letter = self._xls_col_to_letter(col_idx)
                            total_cost_col = {"col": col_letter, "idx": col_idx, "name": str(cell.value).strip()}
        
        result = {}
        if "количество" in column_headers:
            col_idx = column_headers["количество"]["idx"]
            if col_idx < len(row_values) and row_values[col_idx]:
                result["количество"] = {
                    "value": row_values[col_idx],
                    "column": column_headers["количество"]["col"],
                    "name": "Количество"
                }
        
        if cost_unit_col:
            col_idx = cost_unit_col["idx"]
            if col_idx < len(row_values) and row_values[col_idx]:
                result["стоимость_единицы"] = {
                    "value": row_values[col_idx],
                    "column": cost_unit_col["col"],
                    "name": cost_unit_col["name"]
                }
        
        if total_cost_col:
            col_idx = total_cost_col["idx"]
            if col_idx < len(row_values) and row_values[col_idx]:
                result["общая_стоимость"] = {
                    "value": row_values[col_idx],
                    "column": total_cost_col["col"],
                    "name": total_cost_col["name"]
                }
        
        return result

    def _extract_xlsx_row_data(self, file_path: Path, sheet_name: str, row_number: int) -> Dict[str, Any]:
        """Извлечение данных из .xlsx файла."""
        workbook = None
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as error:
            # Если openpyxl не смог открыть, возможно это .xls файл
            if XLRD_AVAILABLE:
                logger.warning(
                    "openpyxl не смог открыть %s для извлечения данных строки (лист %s). "
                    "Пробую через xlrd.",
                    file_path,
                    sheet_name,
                )
                return self._extract_xls_row_data(
                    file_path,
                    sheet_name,
                    row_number,
                    allow_fallback=False,
                )
            logger.error(f"Не удалось открыть документ {file_path}: {error}")
            return {}
        
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            logger.warning(f"Лист '{sheet_name}' не найден в файле {file_path}")
            return {}

        try:
            row_values = []
            max_col = sheet.max_column
            for col_idx in range(1, max_col + 1):
                cell = sheet.cell(row=row_number, column=col_idx)
                row_values.append(self._format_cell_display(cell.value) if cell.value is not None else "")
            
            column_headers = {}
            header_row = None
            
            for header_row_idx in range(1, min(6, sheet.max_row + 1)):
                for col_idx in range(1, min(max_col + 1, 20)):
                    cell = sheet.cell(row=header_row_idx, column=col_idx)
                    if cell.value:
                        header_value = str(cell.value).strip().lower()
                        if header_value == "количество" or (
                            header_value.startswith("количество") and "количество" not in column_headers
                        ):
                            column_headers["количество"] = {"col": cell.column_letter, "idx": col_idx - 1}
                            if header_row is None:
                                header_row = header_row_idx
            
            if "количество" not in column_headers and max_col >= 4:
                column_headers["количество"] = {"col": "D", "idx": 3}
            
            cost_unit_col = None
            total_cost_col = None
            
            for header_row_idx in range(1, min(6, sheet.max_row + 1)):
                for col_idx in range(1, min(max_col + 1, 20)):
                    cell = sheet.cell(row=header_row_idx, column=col_idx)
                    if cell.value:
                        header_value = str(cell.value).strip().lower()
                        if ("стоимость единицы" in header_value or 
                            (header_value.startswith("стоимость") and "единицы" in header_value)):
                            if cost_unit_col is None:
                                cost_unit_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
                        if "общая стоимость" in header_value or header_value.startswith("общая стоимость"):
                            if total_cost_col is None:
                                total_cost_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
            
            result = {}
            if "количество" in column_headers:
                col_idx = column_headers["количество"]["idx"]
                if col_idx < len(row_values) and row_values[col_idx]:
                    result["количество"] = {
                        "value": row_values[col_idx],
                        "column": column_headers["количество"]["col"],
                        "name": "Количество"
                    }
            
            if cost_unit_col:
                col_idx = cost_unit_col["idx"]
                if col_idx < len(row_values) and row_values[col_idx]:
                    result["стоимость_единицы"] = {
                        "value": row_values[col_idx],
                        "column": cost_unit_col["col"],
                        "name": cost_unit_col["name"]
                    }
            elif max_col >= 5:
                for col_idx in [4, 5]:
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["стоимость_единицы"] = {
                            "value": row_values[col_idx],
                            "column": get_column_letter(col_idx + 1),
                            "name": "Стоимость единицы"
                        }
                        break
            
            if total_cost_col:
                col_idx = total_cost_col["idx"]
                if col_idx < len(row_values) and row_values[col_idx]:
                    result["общая_стоимость"] = {
                        "value": row_values[col_idx],
                        "column": total_cost_col["col"],
                        "name": total_cost_col["name"]
                    }
            elif max_col >= 7:
                for col_idx in [6, 7, 8]:
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["общая_стоимость"] = {
                            "value": row_values[col_idx],
                            "column": get_column_letter(col_idx + 1),
                            "name": "Общая стоимость"
                        }
                        break
            
            return result
        finally:
            if workbook:
                workbook.close()

    @staticmethod
    def _xls_col_to_letter(col_idx: int) -> str:
        """Конвертация номера столбца (0-based) в букву Excel (A, B, C, ...)."""
        result = ""
        col_idx += 1
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + (col_idx % 26)) + result
            col_idx //= 26
        return result

    @staticmethod
    def _normalize_cell_value(value: Any) -> Optional[str]:
        """Очистка значения ячейки перед поиском."""
        if isinstance(value, str):
            text = value.strip()
        elif isinstance(value, (int, float)):
            text = f"{value}".strip()
        else:
            text = str(value).strip()

        cleaned = re.sub(r"\s+", " ", text)
        return cleaned.casefold() if cleaned else None

    def _iter_pandas_cells(self, file_path: Path, engine: str = 'xlrd') -> Iterable[Dict[str, Any]]:
        """Итерация по ячейкам через pandas (резервный метод)."""
        if not PANDAS_AVAILABLE:
            raise DocumentSearchError("Библиотека pandas не установлена для резервного открытия файлов.")
        
        # Используем контекстный менеджер для гарантированного закрытия файла
        try:
            with pd.ExcelFile(str(file_path), engine=engine) as excel_file:
                for sheet_name in excel_file.sheet_names:
                    logger.debug(f"Сканирование листа через pandas: {sheet_name}")
                    try:
                        df = pd.read_excel(excel_file, sheet_name=sheet_name, engine=engine, header=None)
                    except Exception as e:
                        logger.warning(f"Не удалось прочитать лист {sheet_name} через pandas: {e}")
                        continue
                    
                    for row_idx, row in df.iterrows():
                        for col_idx, value in enumerate(row):
                            if pd.isna(value) or value == "":
                                continue
                            
                            text = self._normalize_cell_value(value)
                            if text:
                                col_letter = self._xls_col_to_letter(col_idx)
                                cell_address = f"{sheet_name}!{col_letter}{int(row_idx) + 1}"
                                
                                yield {
                                    "text": text,
                                    "display_text": self._format_cell_display(value),
                                    "sheet_name": sheet_name,
                                    "row": int(row_idx) + 1,
                                    "column": col_letter,
                                    "cell_address": cell_address,
                                }
                    # Освобождаем память DataFrame после обработки листа
                    del df
        except Exception as error:
            logger.error(f"Ошибка при открытии файла через pandas (engine={engine}): {error}")
            raise DocumentSearchError(f"Не удалось открыть файл через pandas: {error}") from error
    
    def _extract_full_row_xls(
        self,
        file_path: Path,
        sheet_name: str,
        row_number: int,
        found_column: str,
        context_cols: int
    ) -> Dict[str, Any]:
        """Извлечение полной строки из .xls файла с контекстом."""
        if not XLRD_AVAILABLE:
            return {}
        
        try:
            workbook = xlrd.open_workbook(str(file_path))
        except Exception as error:
            logger.debug(f"Не удалось открыть .xls файл для извлечения полной строки: {error}")
            return {}
        
        try:
            sheet = workbook.sheet_by_name(sheet_name)
        except xlrd.XLRDError:
            logger.debug(f"Лист '{sheet_name}' не найден в файле {file_path}")
            return {}
        
        row_idx = row_number - 1
        if row_idx >= sheet.nrows:
            return {}
        
        # Получаем индекс найденного столбца (0-based для xls)
        found_col_idx = self._letter_to_col_index(found_column) - 1
        
        # Извлекаем заголовки (первые 5 строк)
        column_names = {}
        for header_row_idx in range(min(5, sheet.nrows)):
            for col_idx in range(min(sheet.ncols, 50)):
                cell = sheet.cell(header_row_idx, col_idx)
                if cell.value:
                    col_letter = self._xls_col_to_letter(col_idx)
                    if col_letter not in column_names or header_row_idx == 0:
                        column_names[col_letter] = str(cell.value).strip()
        
        # Извлекаем полную строку
        full_row = []
        for col_idx in range(sheet.ncols):
            cell = sheet.cell(row_idx, col_idx)
            col_letter = self._xls_col_to_letter(col_idx)
            value = self._format_cell_display(cell.value) if cell.value else ""
            full_row.append({
                "column": col_letter,
                "column_name": column_names.get(col_letter, ""),
                "value": value
            })
        
        # Контекст слева и справа
        left_context = []
        right_context = []
        
        start_col = max(0, found_col_idx - context_cols)
        end_col = min(sheet.ncols, found_col_idx + context_cols + 1)
        
        for col_idx in range(start_col, found_col_idx):
            cell = sheet.cell(row_idx, col_idx)
            col_letter = self._xls_col_to_letter(col_idx)
            left_context.append({
                "column": col_letter,
                "column_name": column_names.get(col_letter, ""),
                "value": self._format_cell_display(cell.value) if cell.value else ""
            })
        
        for col_idx in range(found_col_idx + 1, end_col):
            cell = sheet.cell(row_idx, col_idx)
            col_letter = self._xls_col_to_letter(col_idx)
            right_context.append({
                "column": col_letter,
                "column_name": column_names.get(col_letter, ""),
                "value": self._format_cell_display(cell.value) if cell.value else ""
            })
        
        return {
            "full_row": full_row,
            "left_context": left_context,
            "right_context": right_context,
            "column_names": column_names
        }
    
    def _extract_full_row_xlsx(
        self,
        file_path: Path,
        sheet_name: str,
        row_number: int,
        found_column: str,
        context_cols: int
    ) -> Dict[str, Any]:
        """Извлечение полной строки из .xlsx файла с контекстом."""
        workbook = None
        try:
            workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        except Exception as error:
            logger.debug(f"Не удалось открыть .xlsx файл для извлечения полной строки: {error}")
            return {}
        
        try:
            sheet = workbook[sheet_name]
        except KeyError:
            logger.debug(f"Лист '{sheet_name}' не найден в файле {file_path}")
            return {}

        try:
            # Получаем индекс найденного столбца (1-based для xlsx)
            found_col_idx = self._letter_to_col_index(found_column)
            
            # Извлекаем заголовки (первые 5 строк)
            column_names = {}
            for header_row_idx in range(1, min(6, sheet.max_row + 1)):
                for col_idx in range(1, min(sheet.max_column + 1, 50)):
                    cell = sheet.cell(row=header_row_idx, column=col_idx)
                    if cell.value:
                        col_letter = cell.column_letter
                        if col_letter not in column_names or header_row_idx == 1:
                            column_names[col_letter] = str(cell.value).strip()
            
            # Извлекаем полную строку
            full_row = []
            for col_idx in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row_number, column=col_idx)
                col_letter = cell.column_letter
                value = self._format_cell_display(cell.value) if cell.value is not None else ""
                full_row.append({
                    "column": col_letter,
                    "column_name": column_names.get(col_letter, ""),
                    "value": value
                })
            
            # Контекст слева и справа
            left_context = []
            right_context = []
            
            start_col = max(1, found_col_idx - context_cols + 1)
            end_col = min(sheet.max_column + 1, found_col_idx + context_cols + 2)
            
            for col_idx in range(start_col, found_col_idx):
                cell = sheet.cell(row=row_number, column=col_idx)
                left_context.append({
                    "column": cell.column_letter,
                    "column_name": column_names.get(cell.column_letter, ""),
                    "value": self._format_cell_display(cell.value) if cell.value is not None else ""
                })
            
            for col_idx in range(found_col_idx + 1, end_col):
                cell = sheet.cell(row=row_number, column=col_idx)
                right_context.append({
                    "column": cell.column_letter,
                    "column_name": column_names.get(cell.column_letter, ""),
                    "value": self._format_cell_display(cell.value) if cell.value is not None else ""
                })
            
            return {
                "full_row": full_row,
                "left_context": left_context,
                "right_context": right_context,
                "column_names": column_names
            }
        finally:
            if workbook:
                workbook.close()
    
    @staticmethod
    def _letter_to_col_index(column_letter: str) -> int:
        """Конвертация буквы столбца (A, B, C, ...) в индекс (1-based для xlsx, 0-based для xls)."""
        result = 0
        for char in column_letter:
            result = result * 26 + (ord(char.upper()) - ord('A') + 1)
        return result
    
    @staticmethod
    def _format_cell_display(value: Any) -> str:
        """Подготовка значения ячейки для отображения пользователю."""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

