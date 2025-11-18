"""
Сервис для скачивания документации торгов и поиска товаров внутри XLSX-файлов.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple, Union
from concurrent.futures import ThreadPoolExecutor, as_completed
import re
import uuid
import os
import shutil
import subprocess

import requests
from loguru import logger
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from rapidfuzz import fuzz, process
import rarfile
import zipfile
import py7zr
try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False
    logger.warning("Библиотека xlrd не установлена. Поддержка старых .xls файлов будет недоступна.")

from core.database import DatabaseManager
from core.exceptions import DocumentSearchError
from services.helpers.archive_cleanup import ArchiveCleanupManager


class DocumentSearchService:
    """
    Сервис поиска информации в документации торгов.

    Выполняет:
    - Поиск подходящего документа (содержит слово «смета»)
    - Скачивание файла в указанную директорию
    - Парсинг XLSX и нечеткий поиск упоминаний товаров
    """

    KEYWORD_PATTERN = re.compile(r"смет\w*", re.IGNORECASE)
    ARCHIVE_PATTERN = re.compile(
        r"^(?P<base>.+?)(?:[._ -]*(?:part)?(?P<part>\d+))?\.(rar|zip|7z)$",
        re.IGNORECASE,
    )

    DEFAULT_HEADERS = {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
            "(KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36"
        ),
        "Accept": "*/*",
        "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
        "Referer": "https://zakupki.gov.ru/",
        "Connection": "keep-alive",
    }

    def __init__(
        self,
        db_manager: DatabaseManager,
        download_dir: Path,
        unrar_path: Optional[str] = None,
        winrar_path: Optional[str] = None,
        cleanup_manager: Optional[ArchiveCleanupManager] = None,
        progress_callback: Optional[callable] = None,
    ):
        """
        Args:
            db_manager: Менеджер БД с таблицей products
            download_dir: Директория для сохранения файлов
            unrar_path: Путь к инструменту UnRAR (опционально)
            progress_callback: Функция для обновления прогресса (stage, progress, detail)
        """
        self.db_manager = db_manager
        self.download_dir = Path(download_dir)
        self.download_dir.mkdir(parents=True, exist_ok=True)
        self._product_names: Optional[List[str]] = None
        self._keyword = "смета"
        self.http_session = requests.Session()
        self.http_session.headers.update(self.DEFAULT_HEADERS)
        self._rar_tool_configured = False
        self._active_downloads: List[Path] = []
        self._active_extract_dirs: List[Path] = []
        self.cleanup_manager = cleanup_manager or ArchiveCleanupManager()
        self.progress_callback = progress_callback

        effective_unrar = unrar_path or os.environ.get("UNRAR_TOOL")
        self._unrar_path = Path(effective_unrar) if effective_unrar else None

        effective_winrar = winrar_path or os.environ.get("WINRAR_PATH", r"C:\Program Files\WinRAR")
        self._winrar_path = Path(effective_winrar) if effective_winrar else None

        # Установка пути к WinRAR в переменную среды PATH
        if self._winrar_path and self._winrar_path.exists():
            current_path = os.environ.get("PATH", "")
            path_parts = current_path.split(os.pathsep) if current_path else []
            if str(self._winrar_path) not in path_parts:
                os.environ["PATH"] = os.pathsep.join(path_parts + [str(self._winrar_path)]) if path_parts else str(self._winrar_path)
        
        logger.debug(f"DocumentSearchService инициализирован, каталог загрузки: {self.download_dir}")

    def _update_progress(self, stage: str, progress: int, detail: Optional[str] = None):
        """Обновление прогресса через callback"""
        if self.progress_callback:
            try:
                self.progress_callback(stage, progress, detail)
            except Exception as error:
                logger.debug(f"Ошибка при обновлении прогресса: {error}")

    def ensure_products_loaded(self) -> None:
        """Ленивая загрузка названий товаров (по требованию пользователя)."""
        if self._product_names is not None:
            return

        logger.info("Загрузка списка товаров для поиска по документации...")
        query = "SELECT name FROM products WHERE name IS NOT NULL"
        results = self.db_manager.execute_query(query)
        self._product_names = [row.get("name", "").strip() for row in results if row.get("name")]
        logger.info(f"Получено наименований товаров: {len(self._product_names)}")

    def run_document_search(
        self,
        documents: List[Dict[str, Any]],
        tender_id: Optional[int] = None,
        registry_type: str = "44fz",
    ) -> Dict[str, Any]:
        """
        Основной сценарий: найти документы, скачать и выполнить поиск.

        Args:
            documents: Метаданные документов торга

        Returns:
            Словарь с путем к файлу и найденными совпадениями
        """
        if not documents:
            raise DocumentSearchError("У выбранного торга нет приложенных документов.")

        self._update_progress("Подготовка...", 0, "Выбор документов для обработки")
        self._reset_temp_tracking()
        tender_folder = self._prepare_tender_folder(tender_id, registry_type)

        target_docs = self._choose_documents(documents)
        all_downloaded_paths: List[Path] = []
        
        # Группируем документы по архивам, чтобы не скачивать части одного архива несколько раз
        unique_docs_to_download = self._group_documents_by_archive(target_docs, documents)
        logger.info(f"Найдено уникальных документов/архивов для скачивания: {len(unique_docs_to_download)}")
        
        # ЭТАП 1: Скачивание документов
        total_to_download = len(unique_docs_to_download)
        self._update_progress("Скачивание документов", 0, f"Найдено документов для скачивания: {total_to_download}")
        logger.info(f"Начинаю параллельное скачивание {total_to_download} документов/архивов")
        
        downloaded_count = 0
        with ThreadPoolExecutor(max_workers=min(4, total_to_download)) as executor:
            future_to_doc = {
                executor.submit(
                    self._download_required_documents,
                    target_doc,
                    documents,
                    tender_folder,
                ): target_doc
                for target_doc in unique_docs_to_download
            }
            
            for future in as_completed(future_to_doc):
                target_doc = future_to_doc[future]
                doc_name = target_doc.get('file_name') or target_doc.get('document_links', 'unknown')
                try:
                    downloaded_paths = future.result(timeout=300)  # 5 минут на документ
                    all_downloaded_paths.extend(downloaded_paths)
                    downloaded_count += 1
                    # Прогресс от 0% до 30% для этапа скачивания
                    progress = int((downloaded_count / total_to_download) * 30) if total_to_download > 0 else 0
                    self._update_progress(
                        "Скачивание документов",
                        progress,
                        f"Скачано: {downloaded_count}/{total_to_download} - {doc_name}"
                    )
                    logger.info(f"Успешно скачан документ/архив: {doc_name}")
                except Exception as error:
                    logger.error(f"Ошибка при скачивании документа {doc_name}: {error}")
                    # Продолжаем со следующим документом, не прерывая весь процесс
                    continue
        
        if not all_downloaded_paths:
            raise DocumentSearchError("Не удалось скачать ни один документ.")
        
        # Убеждаемся, что прогресс скачивания доходит до 30%
        self._update_progress("Скачивание документов", 30, f"Скачано документов: {downloaded_count}/{total_to_download}")
        
        # ЭТАП 2: Извлечение данных
        self._update_progress("Извлечение данных", 30, "Распаковка архивов и подготовка файлов...")
        workbook_paths = self._prepare_workbook_paths(all_downloaded_paths)
        self._update_progress("Извлечение данных", 60, f"Найдено файлов для обработки: {len(workbook_paths)}")
        
        # ЭТАП 3: Сверка с данными из БД
        self._update_progress("Сверка с данными из БД", 60, "Загрузка списка товаров...")
        matches = self._aggregate_matches_for_workbooks(workbook_paths)

        self._update_progress("Завершено", 100, f"Найдено совпадений: {len(matches)}")
        logger.info(f"Поиск по документации завершен, найдено совпадений: {len(matches)}")

        result = {
            "file_path": str(workbook_paths[0]) if workbook_paths else "",
            "matches": matches,
            "tender_folder": str(tender_folder),
            "downloaded_files": [str(path) for path in all_downloaded_paths],
            "extract_dirs": [str(path) for path in self._active_extract_dirs],
        }

        try:
            self.cleanup_manager.cleanup(
                all_downloaded_paths,
                self._active_extract_dirs,
                matches,
            )
        except Exception as cleanup_error:
            logger.warning(f"Не удалось очистить временные файлы: {cleanup_error}")

        return result

    def _reset_temp_tracking(self) -> None:
        """Сбрасывает списки скачанных файлов и директорий распаковки."""
        self._active_downloads = []
        self._active_extract_dirs = []

    def _prepare_tender_folder(
        self,
        tender_id: Optional[int],
        registry_type: Optional[str],
    ) -> Path:
        """Создает рабочую директорию для файлов торга."""
        if not tender_id:
            fallback_dir = self.download_dir / "tender_temp"
            fallback_dir.mkdir(parents=True, exist_ok=True)
            return fallback_dir

        safe_type = (registry_type or "tender").strip().lower() or "tender"
        folder_name = f"{safe_type}_{tender_id}"
        target_dir = self.download_dir / folder_name
        target_dir.mkdir(parents=True, exist_ok=True)
        return target_dir

    def _register_download(self, path: Path) -> None:
        """Регистрирует скачанный файл для последующей очистки."""
        if path not in self._active_downloads:
            self._active_downloads.append(path)

    def _register_extract_dir(self, path: Path) -> None:
        """Регистрирует директорию распаковки."""
        if path not in self._active_extract_dirs:
            self._active_extract_dirs.append(path)

    def _choose_documents(self, documents: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """
        Выбор документов для обработки:
        - Документы со словом «смета» в названии/ссылке
        - Все Excel файлы (.xlsx, .xls)
        """
        matches: List[Dict[str, Any]] = []
        seen_docs = set()  # Для избежания дубликатов
        
        for doc in documents:
            file_name = (doc.get("file_name") or "").lower()
            link = (doc.get("document_links") or "").lower()
            combined = f"{file_name} {link}"
            
            # Создаем уникальный ключ для документа
            doc_key = (doc.get("file_name"), doc.get("document_links"))
            if doc_key in seen_docs:
                continue
            seen_docs.add(doc_key)
            
            logger.debug(f"Проверка документа: file_name='{file_name}', link='{link}'")
            
            # Проверяем: содержит ли слово "смет" ИЛИ является ли Excel файлом
            is_smeta_doc = self.KEYWORD_PATTERN.search(combined)
            is_excel_file = file_name.endswith((".xlsx", ".xls")) or link.endswith((".xlsx", ".xls"))
            
            if is_smeta_doc or is_excel_file:
                matches.append(doc)
                logger.debug(f"Документ выбран: {doc.get('file_name')} (смета: {bool(is_smeta_doc)}, Excel: {bool(is_excel_file)})")

        if not matches:
            raise DocumentSearchError(
                "Не найден документ со словом «смета» или Excel файл (.xlsx, .xls)."
            )

        # Сортируем по приоритету: сначала Excel файлы, затем архивы
        matches.sort(key=self._document_priority)
        logger.info(f"Выбрано документов для анализа: {len(matches)}")
        for doc in matches:
            logger.info(f"  - {doc.get('file_name') or doc.get('document_links')}")
        return matches

    def _document_priority(self, doc: Dict[str, Any]) -> Tuple[int, str]:
        """Определение приоритета документов (сначала XLSX, затем архивы)."""
        name = (doc.get("file_name") or "").lower()
        if name.endswith(".xlsx"):
            return (0, name)
        if name.endswith(".xls"):
            return (1, name)
        if name.endswith((".rar", ".zip", ".7z")):
            return (2, name)
        return (3, name)
    
    def _group_documents_by_archive(
        self,
        selected_docs: List[Dict[str, Any]],
        all_documents: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """
        Группирует документы по архивам, чтобы не скачивать части одного архива несколько раз.
        
        Для многочастных архивов возвращает только первую часть (остальные будут скачаны автоматически).
        Для обычных документов возвращает их как есть.
        
        Args:
            selected_docs: Выбранные документы для скачивания
            all_documents: Все доступные документы (для поиска связанных частей)
            
        Returns:
            Список уникальных документов для скачивания (по одному на архив)
        """
        unique_docs: List[Dict[str, Any]] = []
        processed_archives: set = set()  # Множество базовых имен обработанных архивов
        
        for doc in selected_docs:
            if not self._is_rar_document(doc):
                # Для не-архивов просто добавляем
                unique_docs.append(doc)
                continue
            
            # Для архивов проверяем, не является ли это частью уже обработанного архива
            base_name, _ = self._split_archive_name(doc.get("file_name"))
            if base_name:
                archive_key = base_name.casefold()
                if archive_key in processed_archives:
                    # Этот архив уже обработан, пропускаем
                    logger.debug(f"Архив {doc.get('file_name')} уже обработан, пропускаем")
                    continue
                processed_archives.add(archive_key)
            
            # Добавляем первую часть архива (остальные части будут скачаны в _download_required_documents)
            unique_docs.append(doc)
        
        logger.info(f"После группировки по архивам: {len(unique_docs)} уникальных документов/архивов")
        return unique_docs

    def _download_required_documents(
        self,
        primary_doc: Dict[str, Any],
        all_documents: List[Dict[str, Any]],
        target_dir: Path,
    ) -> List[Path]:
        """Скачивание основного файла и всех частей архива (если необходимо)."""
        if not self._is_rar_document(primary_doc):
            return [self._download_document(primary_doc, target_dir)]

        related_docs = self._collect_related_archives(primary_doc, all_documents)
        logger.info(
            f"Для архива найдено частей: {len(related_docs)} "
            f"({', '.join(doc.get('file_name') or '' for doc in related_docs)})",
        )
        paths: List[Path] = []
        batch_size = 8
        total_docs = len(related_docs)
        for start in range(0, total_docs, batch_size):
            end = min(start + batch_size, total_docs)
            chunk = related_docs[start:end]
            logger.info(
                f"Скачивание документов {start + 1}-{end} из {total_docs} параллельно",
            )
            chunk_paths = self._download_documents_batch(chunk, target_dir)
            paths.extend(chunk_paths)
        return paths

    def _download_documents_batch(
        self,
        documents: List[Dict[str, Any]],
        target_dir: Path,
    ) -> List[Path]:
        """Параллельная загрузка группы документов с сохранением порядка."""
        ordered_paths: List[Optional[Path]] = [None] * len(documents)
        with ThreadPoolExecutor(max_workers=min(8, len(documents))) as executor:
            future_map = {
                executor.submit(self._download_document, doc, target_dir): index
                for index, doc in enumerate(documents)
            }

            try:
                for future in as_completed(future_map):
                    index = future_map[future]
                    ordered_paths[index] = future.result()
            except Exception as error:
                failed_future = future
                for fut in future_map:
                    fut.cancel()
                failed_index = future_map[failed_future]
                failed_doc = documents[failed_index]
                file_name = failed_doc.get("file_name") or failed_doc.get("document_links")
                raise DocumentSearchError(f"Не удалось скачать документ {file_name}") from error

        return [path for path in ordered_paths if path]

    def _collect_related_archives(
        self,
        primary_doc: Dict[str, Any],
        all_documents: List[Dict[str, Any]],
    ) -> List[Dict[str, Any]]:
        """Сбор всех частей многофайлового архива."""
        primary_base, primary_part = self._split_archive_name(primary_doc.get("file_name"))
        if not primary_base:
            return [primary_doc]

        related: List[Tuple[int, Dict[str, Any]]] = []
        for doc in all_documents:
            if not self._is_rar_document(doc):
                continue
            base, part = self._split_archive_name(doc.get("file_name"))
            if base == primary_base:
                part_index = part if part is not None else 0
                related.append((part_index, doc))

        if not related:
            return [primary_doc]

        related.sort(key=lambda item: item[0])
        ordered_docs: List[Dict[str, Any]] = []
        seen_names = set()
        for _, doc in related:
            name = doc.get("file_name")
            if name in seen_names:
                continue
            seen_names.add(name)
            ordered_docs.append(doc)

        return ordered_docs

    def _is_rar_document(self, document: Dict[str, Any]) -> bool:
        """Проверка, является ли документ архивом (RAR, ZIP, 7Z)."""
        name = (document.get("file_name") or "").lower()
        return name.endswith((".rar", ".zip", ".7z"))
    
    def _is_file_archive(self, file_path: Path) -> bool:
        """
        Проверка, является ли файл архивом по его содержимому (magic bytes).
        
        Это важно для случаев, когда файл имеет расширение Excel, но на самом деле
        является архивом (RAR или 7Z). 
        
        Примечание: .xlsx файлы технически являются ZIP архивами (Office Open XML),
        поэтому мы проверяем только на RAR и 7Z, чтобы не ошибочно определить
        нормальные Excel файлы как архивы.
        
        Args:
            file_path: Путь к файлу для проверки
            
        Returns:
            True если файл является архивом (RAR или 7Z), False иначе
        """
        try:
            if not file_path.exists():
                return False
            
            # Читаем первые байты файла для определения типа
            with open(file_path, 'rb') as f:
                header = f.read(8)
            
            # Magic bytes для архивов:
            # RAR: Rar!\x1a\x07\x00 или Rar!\x1a\x07\x01\x00
            # 7Z: 7z\xbc\xaf\x27\x1c
            
            if len(header) >= 4:
                # RAR файлы начинаются с "Rar!"
                if header[:4] == b'Rar!':
                    return True
                # 7Z файлы начинаются с "7z"
                if len(header) >= 2 and header[:2] == b'7z':
                    return True
            
            return False
        except Exception as error:
            logger.debug(f"Ошибка при проверке файла на архив: {error}")
            # В случае ошибки считаем, что это не архив
            return False

    def _split_archive_name(self, file_name: Optional[str]) -> Tuple[Optional[str], Optional[int]]:
        """Выделение базовой части имени архива и номера части."""
        if not file_name:
            return None, None
        match = self.ARCHIVE_PATTERN.match(file_name)
        if not match:
            return None, None
        base = match.group("base").casefold()
        part = match.group("part")
        return base, int(part) if part is not None else None

    def _prepare_workbook_paths(self, downloaded_paths: List[Path]) -> List[Path]:
        """
        Определение путей ко всем XLSX файлам (напрямую или после распаковки архива).
        
        Обрабатывает все скачанные файлы, проверяя является ли каждый файл архивом
        перед парсингом.
        """
        if not downloaded_paths:
            raise DocumentSearchError("Не удалось скачать документ.")

        all_workbook_paths: List[Path] = []
        
        # Группируем файлы по типам (архивы и Excel)
        archive_paths: List[Path] = []
        excel_paths: List[Path] = []
        
        for path in downloaded_paths:
            suffix = path.suffix.lower()
            if suffix in {".rar", ".zip", ".7z"}:
                archive_paths.append(path)
            elif suffix in {".xlsx", ".xls"}:
                # Проверяем, не является ли файл архивом (маловероятно, но возможно)
                if self._is_file_archive(path):
                    logger.warning(f"Файл {path.name} имеет расширение Excel, но является архивом. Пропускаем.")
                    continue
                excel_paths.append(path)
            else:
                logger.warning(f"Неподдерживаемый формат файла: {path.name} (расширение: {suffix})")
        
        # Обрабатываем архивы
        if archive_paths:
            sorted_archive_paths = sorted(archive_paths, key=self._archive_sort_key)
            # Группируем архивы по базовому имени (для многочастных архивов)
            archive_groups: Dict[str, List[Path]] = {}
            for path in sorted_archive_paths:
                base_name, _ = self._split_archive_name(path.name)
                group_key = base_name or path.stem.casefold()
                if group_key not in archive_groups:
                    archive_groups[group_key] = []
                archive_groups[group_key].append(path)
            
            # Распаковываем каждую группу архивов
            for group_paths in archive_groups.values():
                first_path = group_paths[0]
                suffix = first_path.suffix.lower()
                
                if suffix == ".rar":
                    # Для RAR-архивов используем первый том, unrar сам подцепит остальные части
                    archive_path = first_path
                    if len(group_paths) > 1:
                        logger.info(
                            "Обнаружен многочастный RAR (%s частей). Распаковываю, начиная с %s",
                            len(group_paths),
                            first_path.name,
                        )
                else:
                    archive_path = (
                        self._combine_multi_part_archive(group_paths)
                        if len(group_paths) > 1
                        else first_path
                    )
                extracted_paths = self._extract_archive(archive_path)
                all_workbook_paths.extend(extracted_paths)
        
        # Добавляем Excel файлы напрямую
        all_workbook_paths.extend(excel_paths)
        
        if not all_workbook_paths:
            raise DocumentSearchError("Не найдено ни одного XLSX файла для обработки.")
        
        logger.info(f"Подготовлено {len(all_workbook_paths)} файлов для парсинга")
        return all_workbook_paths

    def debug_process_local_archives(self, archive_paths: List[Union[str, Path]]) -> Dict[str, Any]:
        """
        Тестовая обработка уже скачанных файлов (архивов или XLSX).

        Позволяет быстро проверить распаковку и поиск без повторной загрузки.

        Args:
            archive_paths: Пути к локальным архивам или XLSX файлам

        Returns:
            Результат поиска: путь к файлу и найденные совпадения
        """
        if not archive_paths:
            raise DocumentSearchError("Не переданы пути к локальным файлам.")

        self._reset_temp_tracking()
        resolved_paths: List[Path] = []
        for raw_path in archive_paths:
            path_obj = Path(raw_path).expanduser().resolve()
            if not path_obj.exists():
                raise DocumentSearchError(f"Файл не найден: {path_obj}")
            resolved_paths.append(path_obj)

        logger.info(
            "Запущен тестовый режим обработки локальных файлов: %s",
            ", ".join(path.name for path in resolved_paths),
        )

        workbook_paths = self._prepare_workbook_paths(resolved_paths)
        aggregated_matches = self._aggregate_matches_for_workbooks(workbook_paths)

        logger.info("Тестовая обработка завершена, найдено совпадений: %s", len(aggregated_matches))
        return {
            "workbook_paths": workbook_paths,
            "matches": aggregated_matches,
            "extract_dirs": [str(path) for path in self._active_extract_dirs],
        }

    def _aggregate_matches_for_workbooks(self, workbook_paths: List[Path]) -> List[Dict[str, Any]]:
        """Выполняет поиск по всем XLSX и объединяет совпадения."""
        best_matches: Dict[str, Dict[str, Any]] = {}
        total_files = len(workbook_paths)
        
        for idx, workbook_path in enumerate(workbook_paths):
            # Прогресс от 60% до 95% для этапа сверки с БД
            # Используем (idx + 1) чтобы последний файл давал 95%
            progress = 60 + int(((idx + 1) / total_files) * 35) if total_files > 0 else 60
            file_name = workbook_path.name
            self._update_progress(
                "Сверка с данными из БД",
                progress,
                f"Обработка файла {idx + 1}/{total_files}: {file_name}"
            )
            
            logger.info(f"Поиск по документу: {workbook_path}")
            matches = self._search_workbook_for_products(workbook_path)
            for match in matches:
                product_name = match["product_name"]
                existing = best_matches.get(product_name)
                if existing and existing["score"] >= match["score"]:
                    continue
                best_matches[product_name] = {
                    **match,
                    "source_file": str(workbook_path),
                }

        # Убеждаемся, что прогресс доходит до 95% после обработки всех файлов
        self._update_progress("Сверка с данными из БД", 95, f"Обработка завершена, найдено совпадений: {len(best_matches)}")
        
        sorted_matches = sorted(best_matches.values(), key=lambda item: item["score"], reverse=True)
        return sorted_matches[:50]

    def _combine_multi_part_archive(self, archive_paths: List[Path]) -> Path:
        """
        Склейка многофайлового архива в один файл.
        
        Для архивов, разбитых на части (part1.rar, part2.rar и т.д.),
        склеивает все части в один файл перед распаковкой.
        
        Args:
            archive_paths: Список путей к частям архива в правильном порядке
            
        Returns:
            Путь к склеенному архиву
        """
        if len(archive_paths) == 1:
            return archive_paths[0]
        
        logger.info(f"Обнаружен многофайловый архив из {len(archive_paths)} частей. Начинаю склейку...")
        
        # Определяем базовое имя и расширение
        first_path = archive_paths[0]
        base_name, _ = self._split_archive_name(first_path.name)
        
        if not base_name:
            # Если не удалось определить базовое имя, используем имя первого файла без номера части
            base_name = first_path.stem.rsplit('.part', 1)[0] if '.part' in first_path.stem.lower() else first_path.stem
        
        # Создаем имя для склеенного архива
        suffix = first_path.suffix.lower()
        combined_path = first_path.parent / f"{base_name}_combined{suffix}"
        
        # Склеиваем все части в один файл
        try:
            with open(combined_path, 'wb') as combined_file:
                for part_path in sorted(archive_paths, key=lambda p: self._get_part_number(p)):
                    logger.debug(f"Добавляю часть: {part_path.name}")
                    with open(part_path, 'rb') as part_file:
                        # Копируем данные по частям для экономии памяти
                        while True:
                            chunk = part_file.read(8192)
                            if not chunk:
                                break
                            combined_file.write(chunk)
            
            logger.info(f"Многофайловый архив успешно склеен: {combined_path.name}")
            return combined_path
            
        except Exception as error:
            logger.error(f"Ошибка при склейке архива: {error}")
            # Если склейка не удалась, пробуем использовать первую часть
            logger.warning(f"Использую первую часть архива: {first_path.name}")
            return first_path
    
    def _get_part_number(self, path: Path) -> int:
        """Извлечение номера части из имени файла."""
        match = self.ARCHIVE_PATTERN.match(path.name)
        if match and match.group("part"):
            return int(match.group("part"))
        # Если номер части не найден, проверяем другие форматы
        name_lower = path.name.lower()
        if '.part' in name_lower:
            # Пытаемся извлечь номер из формата .part01, .part1 и т.д.
            part_match = re.search(r'\.part(\d+)', name_lower)
            if part_match:
                return int(part_match.group(1))
        return 0  # Если номер не найден, считаем первой частью

    def _archive_sort_key(self, path: Path) -> Tuple[str, int]:
        """Ключ сортировки для частей архивов."""
        base_name, _ = self._split_archive_name(path.name)
        base = base_name or path.stem.casefold()
        return base, self._get_part_number(path)

    def _extract_archive(self, archive_path: Path) -> List[Path]:
        """Распаковка архива (ZIP, RAR, 7Z) и поиск всех XLSX внутри."""
        logger.info(f"Распаковка архива {archive_path.name}")
        extract_dir = archive_path.parent / f"extract_{archive_path.stem}_{uuid.uuid4().hex[:6]}"
        extract_dir.mkdir(parents=True, exist_ok=True)
        self._register_extract_dir(extract_dir)
        suffix = archive_path.suffix.lower()

        try:
            if suffix == ".zip":
                self._extract_zip_archive(archive_path, extract_dir)
            elif suffix == ".rar":
                self._extract_rar_archive(archive_path, extract_dir)
            elif suffix == ".7z":
                self._extract_7z_archive(archive_path, extract_dir)
            else:
                raise DocumentSearchError(f"Неподдерживаемый формат архива: {suffix}")
        except Exception as error:
            logger.error(f"Ошибка распаковки архива {archive_path.name}: {error}")
            raise DocumentSearchError(f"Ошибка распаковки архива: {error}") from error

        xlsx_files = [
            path for path in extract_dir.rglob("*.xlsx")
            if not path.name.startswith("~$")
        ]
        if not xlsx_files:
            raise DocumentSearchError("В архиве не найден ни один XLSX файл.")

        logger.info(f"Найдено XLSX после распаковки: {len(xlsx_files)}")
        for file in xlsx_files:
            logger.debug(f"  - {file}")
        return xlsx_files

    def _extract_zip_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка ZIP-архива с правильной обработкой кодировок."""
        logger.debug(f"Распаковка ZIP-архива: {archive_path.name}")
        try:
            with zipfile.ZipFile(archive_path, 'r') as zip_ref:
                for zip_info in zip_ref.infolist():
                    # Попытка декодировать имена файлов с использованием cp866
                    try:
                        zip_info.filename = zip_info.filename.encode('cp437').decode('cp866')
                    except (UnicodeEncodeError, UnicodeDecodeError):
                        # Если не удалось декодировать, оставляем оригинальное имя
                        pass
                    zip_ref.extract(zip_info, extract_dir)
            logger.info(f"ZIP-архив {archive_path.name} успешно распакован")
        except zipfile.BadZipFile as error:
            logger.error(f"Поврежденный ZIP-архив: {error}")
            raise DocumentSearchError("ZIP-архив поврежден или имеет неподдерживаемый формат.") from error
        except Exception as error:
            logger.error(f"Ошибка распаковки ZIP-архива: {error}")
            raise DocumentSearchError(f"Ошибка распаковки ZIP-архива: {error}") from error

    def _extract_rar_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка RAR-архива."""
        logger.debug(f"Распаковка RAR-архива: {archive_path.name}")
        unrar_executable = self._ensure_unrar_available()
        logger.debug(f"Используемый UNRAR_TOOL: {unrar_executable}")
        logger.debug(f"Содержимое директории распаковки: {list(extract_dir.iterdir()) if extract_dir.exists() else 'пусто'}")
        logger.debug(f"Размер архива: {archive_path.stat().st_size} байт")

        command = [
            unrar_executable,
            "x",  # extract with full paths
            "-y",  # assume Yes on all queries
            archive_path.name,
            str(extract_dir),
        ]

        logger.debug(f"Запускаю распаковку: {' '.join(command)} (cwd={archive_path.parent})")
        result = subprocess.run(
            command,
            cwd=str(archive_path.parent),
            capture_output=True,
            text=True,
        )

        if result.returncode != 0:
            logger.error("Не удалось распаковать RAR-архив. Код завершения: %s", result.returncode)
            logger.error("STDOUT:\n%s", result.stdout)
            logger.error("STDERR:\n%s", result.stderr)
            self._log_rar_volume_debug_info(archive_path)
            raise DocumentSearchError("RAR-архив поврежден или имеет неподдерживаемый формат.")

        logger.debug("STDOUT:\n%s", result.stdout)
        logger.info(f"RAR-архив {archive_path.name} успешно распакован")

    def _log_rar_volume_debug_info(self, archive_path: Path) -> None:
        """Логирование отладочной информации при ошибках распаковки."""
        logger.debug("=== Отладочная информация по архиву ===")
        logger.debug(f"Имя архива: {archive_path.name}")
        logger.debug(f"Расположение: {archive_path.parent}")
        logger.debug(f"Размер файла: {archive_path.stat().st_size} байт")
        logger.debug(f"Расширение: {archive_path.suffix}")

        # Логируем соседние файлы (вдруг требуется другой том)
        base_name = archive_path.stem.replace("_combined", "")
        for neighbor in archive_path.parent.glob(f"{base_name}*"):
            if neighbor == archive_path:
                continue
            logger.debug(f"Найден соседний файл: {neighbor.name}")

    def _extract_7z_archive(self, archive_path: Path, extract_dir: Path) -> None:
        """Распаковка 7Z-архива."""
        logger.debug(f"Распаковка 7Z-архива: {archive_path.name}")
        try:
            with py7zr.SevenZipFile(archive_path, mode='r') as archive:
                archive.extractall(path=extract_dir)
            logger.info(f"7Z-архив {archive_path.name} успешно распакован")
        except Exception as error:
            logger.error(f"Ошибка распаковки 7Z-архива {archive_path.name}: {error}")
            raise DocumentSearchError(f"Ошибка распаковки 7Z-архива: {error}") from error

    def _ensure_unrar_available(self) -> None:
        """Проверка наличия инструмента распаковки RAR."""
        if self._rar_tool_configured and rarfile.UNRAR_TOOL:
            return rarfile.UNRAR_TOOL

        env_tool = os.environ.get("UNRAR_TOOL")
        candidate_paths = self._collect_unrar_candidates(env_tool)

        for candidate in candidate_paths:
            if not candidate:
                continue
            candidate_path = Path(candidate)
            if candidate_path.exists():
                rarfile.UNRAR_TOOL = str(candidate_path)
                self._rar_tool_configured = True
                logger.info(f"Найден инструмент распаковки RAR: {candidate_path}")
                return rarfile.UNRAR_TOOL

        self._rar_tool_configured = False
        logger.warning(
            "Инструмент распаковки RAR не найден. Установите WinRAR/UnRAR "
            "и задайте путь через переменную окружения UNRAR_TOOL.",
        )
        raise DocumentSearchError(
            "Инструмент для распаковки RAR не найден. "
            "Установите WinRAR/UnRAR и задайте UNRAR_TOOL.",
        )

    def _collect_unrar_candidates(self, env_tool: Optional[str]) -> List[str]:
        """Собирает список потенциальных путей к инструментам распаковки RAR."""
        candidates: List[str] = []

        # 1. Прямые пути
        for direct in (self._unrar_path,):
            if direct:
                candidates.append(str(direct))

        if env_tool:
            candidates.append(env_tool)

        # 2. Пути из настроек WINRAR_PATH + популярные exe имена
        possible_dirs = [
            str(self._winrar_path) if self._winrar_path else None,
            os.environ.get("WINRAR_PATH"),
            r"C:\Program Files\WinRAR",
            r"C:\Program Files (x86)\WinRAR",
            r"C:\Program Files\UNRAR",
            r"C:\Program Files (x86)\UNRAR",
        ]

        exe_names = ["UnRAR.exe", "UNRAR.exe", "UNRARG.exe", "WinRAR.exe", "rar.exe"]
        for directory in possible_dirs:
            if not directory:
                continue
            for exe_name in exe_names:
                candidates.append(os.path.join(directory, exe_name))

        # 3. Проверяем, можно ли найти exe через PATH
        for exe_name in exe_names:
            which_path = shutil.which(exe_name)
            if which_path:
                candidates.append(which_path)

        # 4. Удаляем дубликаты, сохраняя порядок
        seen = set()
        unique_candidates = []
        for path in candidates:
            if path and path not in seen:
                unique_candidates.append(path)
                seen.add(path)

        return unique_candidates

    def _download_document(
        self,
        document: Dict[str, Any],
        target_dir: Optional[Path] = None,
    ) -> Path:
        """Скачивание XLSX-файла по ссылке."""
        url = document.get("document_links")
        if not url:
            raise DocumentSearchError("У выбранного документа отсутствует ссылка для скачивания.")

        raw_name = document.get("file_name") or f"document_{document.get('id', 'unknown')}"
        file_name = self._sanitize_filename(raw_name)
        suffix = Path(file_name).suffix.lower()
        if not suffix:
            file_name = f"{file_name}.xlsx"

        destination_dir = target_dir or self.download_dir
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = destination_dir / file_name
        logger.info(f"Начинаю скачивание документа '{file_name}' по ссылке {url}")

        try:
            # Таймауты: (connect_timeout, read_timeout)
            # connect_timeout - время на установку соединения
            # read_timeout - время на чтение данных (важно для больших файлов)
            timeout_config = (30, 300)  # 30 сек на подключение, 5 минут на чтение
            
            response = self.http_session.get(
                url,
                timeout=timeout_config,
                stream=True,
                allow_redirects=True
            )
            response.raise_for_status()
            
            # Получаем размер файла для прогресса (если доступен)
            total_size = response.headers.get('content-length')
            if total_size:
                total_size = int(total_size)
                logger.debug(f"Размер файла: {total_size / 1024 / 1024:.2f} MB")
            
            downloaded = 0
            last_progress_update = 0
            with destination.open("wb") as file:
                for chunk in response.iter_content(chunk_size=8192):
                    if chunk:
                        file.write(chunk)
                        downloaded += len(chunk)
                        # Обновляем прогресс каждые 1MB или каждые 5% (что быстрее)
                        if total_size:
                            progress_percent = (downloaded / total_size) * 100
                            # Обновляем каждые 1MB или каждые 5%
                            if (downloaded - last_progress_update >= 1024 * 1024) or (progress_percent - last_progress_update >= 5):
                                logger.debug(f"Прогресс скачивания {file_name}: {progress_percent:.1f}%")
                                last_progress_update = progress_percent
                        else:
                            # Если размер неизвестен, обновляем каждые 1MB
                            if downloaded - last_progress_update >= 1024 * 1024:
                                logger.debug(f"Скачано {file_name}: {downloaded / 1024 / 1024:.1f} MB")
                                last_progress_update = downloaded
                            
        except requests.Timeout as error:
            logger.error(f"Таймаут при скачивании документа '{file_name}': {error}")
            # Удаляем частично скачанный файл
            if destination.exists():
                try:
                    destination.unlink()
                except Exception:
                    pass
            raise DocumentSearchError(f"Таймаут при скачивании документа '{file_name}'. Файл слишком большой или соединение медленное.") from error
        except requests.RequestException as error:
            logger.error(f"Ошибка скачивания документа '{file_name}': {error}")
            # Удаляем частично скачанный файл
            if destination.exists():
                try:
                    destination.unlink()
                except Exception:
                    pass
            raise DocumentSearchError(f"Не удалось скачать документ '{file_name}'.") from error

        logger.info(f"Документ сохранен: {destination}")
        self._register_download(destination)
        return destination

    def _search_workbook_for_products(self, file_path: Path) -> List[Dict[str, Any]]:
        """Парсинг XLSX и поиск совпадений с названиями товаров по ключевым словам."""
        self.ensure_products_loaded()
        if not self._product_names:
            logger.warning("Список товаров пуст, поиск не будет выполнен.")
            return []

        # Подготавливаем паттерны поиска для каждого товара
        product_patterns: Dict[str, Dict[str, Any]] = {}
        for product_name in self._product_names:
            pattern = self._extract_keywords(product_name)
            if pattern:
                product_patterns[product_name] = pattern

        found_matches: Dict[str, Dict[str, Any]] = {}

        for cell_info in self._iter_workbook_cells(file_path):
            text = cell_info["text"]
            display_text = cell_info["display_text"]
            cell_matches: List[Tuple[str, Dict[str, Any]]] = []

            # Проверяем каждый товар на наличие его ключевых слов в тексте
            for product_name, pattern in product_patterns.items():
                match_result = self._check_keywords_match(text, pattern, product_name)
                if match_result["found"]:
                    cell_matches.append((product_name, match_result))

            if not cell_matches:
                continue

            # Если в ячейке найдено полное совпадение хотя бы для одного товара,
            # отдаем предпочтение только этим совпадениям
            full_matches = [
                (product_name, match_result)
                for product_name, match_result in cell_matches
                if match_result["full_match"]
            ]
            if full_matches:
                matches_to_apply = full_matches
            else:
                matches_to_apply = cell_matches

            # Выбираем лучшее совпадение в ячейке (без дублей)
            best_product, best_match = max(matches_to_apply, key=lambda item: item[1]["score"])

            # Если товар уже найден, обновляем только если нашли более точное совпадение
            existing = found_matches.get(best_product)
            if existing and existing.get("score", 0) >= best_match["score"]:
                continue

            # Извлекаем данные из строки Excel
            row_data = self._extract_row_data(
                file_path,
                cell_info["sheet_name"],
                cell_info["row"]
            )
            
            found_matches[best_product] = {
                "product_name": best_product,
                "score": best_match["score"],
                "matched_text": text,
                "matched_display_text": display_text,
                "sheet_name": cell_info["sheet_name"],
                "row": cell_info["row"],
                "column": cell_info["column"],
                "cell_address": cell_info["cell_address"],
                "matched_keywords": best_match["matched_keywords"],
                "row_data": row_data,
            }

        sorted_matches = sorted(found_matches.values(), key=lambda item: item["score"], reverse=True)
        return sorted_matches[:50]

    def _extract_keywords(self, product_name: str) -> Optional[Dict[str, Any]]:
        """
        Извлечение ключевых слов из названия товара.
        
        Примеры:
        - "ДенсТоп ЭП 203 (Комплект_1)" -> ["денстоп", "эп", "203"]
        - "Реолен Адмикс Плюс" -> ["реолен", "адмикс", "плюс"]
        - "Маногард 133 Фер (Комплект_2)" -> ["маногард", "133", "фер"]
        """
        # Убираем скобки и их содержимое
        cleaned = re.sub(r'\([^)]*\)', '', product_name).strip()
        if not cleaned:
            return []

        keywords: List[str] = []

        # Добавляем полное название (без скобок) как ключевую фразу
        normalized_full_name = re.sub(r'\s+', ' ', cleaned).casefold()
        full_phrase = normalized_full_name if len(normalized_full_name) >= 3 else None

        # Разбиваем на слова (учитываем кириллицу, латиницу, цифры)
        words = re.findall(r'[а-яёА-ЯЁa-zA-Z0-9]+', cleaned, re.IGNORECASE)

        for word in words:
            word_lower = word.casefold()
            # Пропускаем слишком короткие слова и аббревиатуры (менее 3 символов)
            if len(word_lower) < 3:
                continue
            # Пропускаем слова, состоящие только из цифр (не ищем по «205», «470» и т.п.)
            if word_lower.isdigit():
                continue
            keywords.append(word_lower)

        # Удаляем дубликаты, сохраняя порядок
        seen = set()
        unique_keywords = []
        for keyword in keywords:
            if keyword not in seen:
                unique_keywords.append(keyword)
                seen.add(keyword)

        if not keywords and not full_phrase:
            return None

        return {
            "full_phrase": full_phrase,
            "tokens": unique_keywords,
        }

    def _check_keywords_match(self, text: str, pattern: Dict[str, Any], product_name: str) -> Dict[str, Any]:
        """
        Проверка наличия ключевых слов в тексте с учетом возможных опечаток.
        
        Returns:
            Dict с полями:
            - found: bool - найдены ли ключевые слова
            - score: float - процент совпадения (100.0 для точного, 85.0+ для хорошего, 35.0+ для слабого)
            - matched_keywords: List[str] - список найденных ключевых слов
        """
        keywords = pattern.get("tokens", [])
        full_phrase = pattern.get("full_phrase")

        if not keywords and not full_phrase:
            return {"found": False, "score": 0.0, "matched_keywords": [], "full_match": False}
        
        text_lower = text.casefold()
        matched_keywords = []

        # ШАГ 1: Проверяем точное совпадение полного названия товара (без учета регистра)
        # Убираем скобки из названия для сравнения
        product_name_clean = re.sub(r'\([^)]*\)', '', product_name).strip()
        product_name_normalized = re.sub(r'\s+', ' ', product_name_clean).casefold()
        
        if product_name_normalized and product_name_normalized in text_lower:
            return {
                "found": True,
                "score": 100.0,
                "matched_keywords": [product_name_normalized],
                "full_match": True,
            }

        # ШАГ 2: Проверяем полное совпадение фразы (без скобок)
        if full_phrase and full_phrase in text_lower:
            matched_keywords.append(full_phrase)
            return {
                "found": True,
                "score": 100.0,
                "matched_keywords": matched_keywords,
                "full_match": True,
            }
        
        # ШАГ 3: Если точного совпадения нет, ищем по ключевым словам
        for keyword in keywords:
            # Сначала проверяем точное вхождение (без учета регистра)
            if keyword in text_lower:
                matched_keywords.append(keyword)
                continue
            
            # Для коротких слов (3+ символа) проверяем опечатки с помощью fuzzy matching
            if len(keyword) >= 3:
                # Ищем слово в тексте с учетом опечаток
                words_in_text = re.findall(r'[а-яёА-ЯЁa-zA-Z0-9]+', text_lower)
                for word in words_in_text:
                    if len(word) >= 3:
                        similarity = fuzz.ratio(keyword, word)
                        # Порог 85% для учета опечаток (например, "реолен" vs "риолен")
                        if similarity >= 85:
                            matched_keywords.append(keyword)
                            break
        
        if not matched_keywords:
            return {"found": False, "score": 0.0, "matched_keywords": [], "full_match": False}
        
        # Считаем процент найденных ключевых слов
        keywords_ratio = len(matched_keywords) / len(keywords) if keywords else 0
        
        # Определяем score по градации:
        # - 85%+ если найдено большинство ключевых слов (>= 60% ключевых слов)
        # - 35%+ если найдено только часть ключевых слов (>= 30% ключевых слов)
        # - иначе 0 (не выводим)
        if keywords_ratio >= 0.6:
            score = 85.0 + (keywords_ratio - 0.6) * 15.0  # 85-100 для хороших совпадений
        elif keywords_ratio >= 0.3:
            score = 35.0 + (keywords_ratio - 0.3) * 50.0  # 35-85 для слабых совпадений
        else:
            return {"found": False, "score": 0.0, "matched_keywords": [], "full_match": False}
        
        found = len(matched_keywords) > 0
        
        return {
            "found": found,
            "score": score,
            "matched_keywords": matched_keywords,
            "full_match": False,
        }

    def _extract_row_data(self, file_path: Path, sheet_name: str, row_number: int) -> Dict[str, Any]:
        """
        Извлекает данные из строки Excel, включая столбец "Количество" и другие столбцы.
        
        Поддерживает оба формата:
        - .xlsx через openpyxl
        - .xls через xlrd
        
        Args:
            file_path: Путь к файлу Excel
            sheet_name: Имя листа
            row_number: Номер строки (1-based)
            
        Returns:
            Dict с данными строки: количество, стоимость единицы, общая стоимость и т.д.
        """
        suffix = file_path.suffix.lower()
        
        try:
            if suffix == ".xls":
                # Обработка старого формата .xls через xlrd
                if not XLRD_AVAILABLE:
                    logger.warning("xlrd не доступен, пропускаю извлечение данных для .xls файла")
                    return {}
                
                workbook = xlrd.open_workbook(str(file_path))
                sheet = workbook.sheet_by_name(sheet_name)
                
                # Читаем строку данных (xlrd использует 0-based индексы)
                row_idx = row_number - 1
                if row_idx >= sheet.nrows:
                    return {}
                
                row_values = []
                max_col = sheet.ncols
                for col_idx in range(max_col):
                    cell = sheet.cell(row_idx, col_idx)
                    row_values.append(self._format_cell_display(cell.value) if cell.value else "")
                
                # Ищем заголовки столбцов
                column_headers = {}
                header_row = None
                
                for header_row_idx in range(min(5, sheet.nrows)):
                    for col_idx in range(min(max_col, 20)):
                        cell = sheet.cell(header_row_idx, col_idx)
                        if cell.value:
                            header_value = str(cell.value).strip().lower()
                            if header_value == "количество" or (header_value.startswith("количество") and "количество" not in column_headers):
                                col_letter = self._xls_col_to_letter(col_idx)
                                column_headers["количество"] = {"col": col_letter, "idx": col_idx}
                                if header_row is None:
                                    header_row = header_row_idx
                
                if "количество" not in column_headers and max_col >= 4:
                    column_headers["количество"] = {"col": "D", "idx": 3}
                
                # Ищем столбцы стоимости
                cost_unit_col = None
                total_cost_col = None
                
                for header_row_idx in range(min(5, sheet.nrows)):
                    for col_idx in range(min(max_col, 20)):
                        cell = sheet.cell(header_row_idx, col_idx)
                        if cell.value:
                            header_value = str(cell.value).strip().lower()
                            if ("стоимость единицы" in header_value or 
                                (header_value.startswith("стоимость") and "единицы" in header_value)):
                                if cost_unit_col is None:
                                    col_letter = self._xls_col_to_letter(col_idx)
                                    cost_unit_col = {"col": col_letter, "idx": col_idx, "name": str(cell.value).strip()}
                            if "общая стоимость" in header_value or header_value.startswith("общая стоимость"):
                                if total_cost_col is None:
                                    col_letter = self._xls_col_to_letter(col_idx)
                                    total_cost_col = {"col": col_letter, "idx": col_idx, "name": str(cell.value).strip()}
                
                # Извлекаем значения
                result = {}
                if "количество" in column_headers:
                    col_idx = column_headers["количество"]["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["количество"] = {
                            "value": row_values[col_idx],
                            "column": column_headers["количество"]["col"],
                            "name": "Количество"
                        }
                
                if cost_unit_col:
                    col_idx = cost_unit_col["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["стоимость_единицы"] = {
                            "value": row_values[col_idx],
                            "column": cost_unit_col["col"],
                            "name": cost_unit_col["name"]
                        }
                
                if total_cost_col:
                    col_idx = total_cost_col["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["общая_стоимость"] = {
                            "value": row_values[col_idx],
                            "column": total_cost_col["col"],
                            "name": total_cost_col["name"]
                        }
                
                return result
            else:
                # Обработка нового формата .xlsx через openpyxl
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook[sheet_name]
                
                # Читаем строку данных
                row_values = []
                max_col = sheet.max_column
                for col_idx in range(1, max_col + 1):
                    cell = sheet.cell(row=row_number, column=col_idx)
                    row_values.append(self._format_cell_display(cell.value) if cell.value is not None else "")
                
                # Ищем заголовки столбцов (проверяем первые 5 строк для многоуровневых заголовков)
                column_headers = {}
                header_row = None
                
                # Сначала ищем точное совпадение "Количество" (может быть в разных строках)
                for header_row_idx in range(1, min(6, sheet.max_row + 1)):
                    for col_idx in range(1, min(max_col + 1, 20)):  # Ограничиваем поиск первыми 20 столбцами
                        cell = sheet.cell(row=header_row_idx, column=col_idx)
                        if cell.value:
                            header_value = str(cell.value).strip().lower()
                            
                            # Ищем столбец "Количество" (точное совпадение или содержит)
                            if header_value == "количество" or (header_value.startswith("количество") and "количество" not in column_headers):
                                column_headers["количество"] = {"col": cell.column_letter, "idx": col_idx - 1}
                                if header_row is None:
                                    header_row = header_row_idx
                    
                    # Также проверяем объединенные ячейки и соседние ячейки
                    if "количество" not in column_headers:
                        # Ищем в тексте ячеек, которые могут содержать "количество" как часть
                        for col_idx in range(1, min(max_col + 1, 20)):
                            cell = sheet.cell(row=header_row_idx, column=col_idx)
                            if cell.value:
                                header_value = str(cell.value).strip().lower()
                                # Проверяем, не является ли это частью многоуровневого заголовка
                                if "количество" in header_value and "количество" not in column_headers:
                                    # Проверяем, что это не часть другого слова
                                    if header_value == "количество" or header_value.startswith("количество"):
                                        column_headers["количество"] = {"col": cell.column_letter, "idx": col_idx - 1}
                                        if header_row is None:
                                            header_row = header_row_idx
                                        break
                
                # Если не нашли "Количество" по тексту, пробуем найти по позиции (обычно это 4-й столбец)
                if "количество" not in column_headers and max_col >= 4:
                    # Проверяем, может быть столбец D (4-й) - это часто "Количество"
                    column_headers["количество"] = {"col": "D", "idx": 3}
                
                # Ищем названия столбцов для стоимости
                cost_unit_col = None
                total_cost_col = None
                
                for header_row_idx in range(1, min(6, sheet.max_row + 1)):
                    for col_idx in range(1, min(max_col + 1, 20)):
                        cell = sheet.cell(row=header_row_idx, column=col_idx)
                        if cell.value:
                            header_value = str(cell.value).strip().lower()
                            # Ищем "стоимость единицы"
                            if ("стоимость единицы" in header_value or 
                                (header_value.startswith("стоимость") and "единицы" in header_value)):
                                if cost_unit_col is None:
                                    cost_unit_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
                            # Ищем "общая стоимость"
                            if "общая стоимость" in header_value or header_value.startswith("общая стоимость"):
                                if total_cost_col is None:
                                    total_cost_col = {"col": cell.column_letter, "idx": col_idx - 1, "name": str(cell.value).strip()}
                
                workbook.close()
                
                # Извлекаем значения из нужных столбцов
                result = {}
                if "количество" in column_headers:
                    col_idx = column_headers["количество"]["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["количество"] = {
                            "value": row_values[col_idx],
                            "column": column_headers["количество"]["col"],
                            "name": "Количество"
                        }
                
                # Столбец со стоимостью единицы
                if cost_unit_col:
                    col_idx = cost_unit_col["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["стоимость_единицы"] = {
                            "value": row_values[col_idx],
                            "column": cost_unit_col["col"],
                            "name": cost_unit_col["name"]
                        }
                elif max_col >= 5:
                    # Fallback: пробуем столбцы E, F
                    for col_idx in [4, 5]:
                        if col_idx < len(row_values) and row_values[col_idx]:
                            result["стоимость_единицы"] = {
                                "value": row_values[col_idx],
                                "column": get_column_letter(col_idx + 1),
                                "name": "Стоимость единицы"
                            }
                            break
                
                # Столбец с общей стоимостью
                if total_cost_col:
                    col_idx = total_cost_col["idx"]
                    if col_idx < len(row_values) and row_values[col_idx]:
                        result["общая_стоимость"] = {
                            "value": row_values[col_idx],
                            "column": total_cost_col["col"],
                            "name": total_cost_col["name"]
                        }
                elif max_col >= 7:
                    # Fallback: пробуем столбцы G, H, I
                    for col_idx in [6, 7, 8]:
                        if col_idx < len(row_values) and row_values[col_idx]:
                            result["общая_стоимость"] = {
                                "value": row_values[col_idx],
                                "column": get_column_letter(col_idx + 1),
                                "name": "Общая стоимость"
                            }
                            break
                
                return result
            
        except Exception as error:
            logger.debug(f"Ошибка при извлечении данных строки {row_number} из {sheet_name}: {error}")
            return {}

    def _iter_workbook_cells(self, file_path: Path) -> Iterable[Dict[str, Any]]:
        """
        Итерация по всем ячейкам Excel (XLSX или XLS) с информацией о позиции.
        
        Поддерживает оба формата:
        - .xlsx через openpyxl
        - .xls через xlrd
        """
        suffix = file_path.suffix.lower()
        
        if suffix == ".xls":
            # Обработка старого формата .xls через xlrd
            if not XLRD_AVAILABLE:
                raise DocumentSearchError(
                    "Библиотека xlrd не установлена. Установите её для поддержки .xls файлов: pip install xlrd"
                )
            
            try:
                workbook = xlrd.open_workbook(str(file_path))
            except Exception as error:
                logger.error(f"Не удалось открыть документ {file_path}: {error}")
                raise DocumentSearchError("Поврежденный или неподдерживаемый XLS файл.") from error
            
            for sheet in workbook.sheets():
                logger.debug(f"Сканирование листа: {sheet.name}")
                for row_idx in range(sheet.nrows):
                    for col_idx in range(sheet.ncols):
                        cell = sheet.cell(row_idx, col_idx)
                        if cell.value is None or cell.value == "":
                            continue
                        
                        text = self._normalize_cell_value(cell.value)
                        if text:
                            # Конвертируем номер столбца в букву (A, B, C, ...)
                            col_letter = self._xls_col_to_letter(col_idx)
                            cell_address = f"{sheet.name}!{col_letter}{row_idx + 1}"
                            
                            yield {
                                "text": text,
                                "display_text": self._format_cell_display(cell.value),
                                "sheet_name": sheet.name,
                                "row": row_idx + 1,  # xlrd использует 0-based, конвертируем в 1-based
                                "column": col_letter,
                                "cell_address": cell_address,
                            }
        else:
            # Обработка нового формата .xlsx через openpyxl
            try:
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
            except Exception as error:
                logger.error(f"Не удалось открыть документ {file_path}: {error}")
                raise DocumentSearchError("Поврежденный или неподдерживаемый XLSX файл.") from error

            for sheet in workbook.worksheets:
                logger.debug(f"Сканирование листа: {sheet.title}")
                for row_idx, row in enumerate(sheet.iter_rows(values_only=False), start=1):
                    for cell in row:
                        if cell.value is None:
                            continue
                        text = self._normalize_cell_value(cell.value)
                        if text:
                            yield {
                                "text": text,
                                "display_text": self._format_cell_display(cell.value),
                                "sheet_name": sheet.title,
                                "row": row_idx,
                                "column": cell.column_letter,
                                "cell_address": f"{sheet.title}!{cell.coordinate}",
                            }

            workbook.close()
    
    @staticmethod
    def _xls_col_to_letter(col_idx: int) -> str:
        """Конвертация номера столбца (0-based) в букву Excel (A, B, C, ...)."""
        result = ""
        col_idx += 1  # Конвертируем в 1-based
        while col_idx > 0:
            col_idx -= 1
            result = chr(65 + (col_idx % 26)) + result
            col_idx //= 26
        return result

    @staticmethod
    def _normalize_cell_value(value: Any) -> Optional[str]:
        """Очистка значения ячейки перед поиском."""
        if isinstance(value, str):
            text = value.strip()
        elif isinstance(value, (int, float)):
            text = f"{value}".strip()
        else:
            text = str(value).strip()

        cleaned = re.sub(r"\s+", " ", text)
        return cleaned.casefold() if cleaned else None

    @staticmethod
    def _format_cell_display(value: Any) -> str:
        """Подготовка значения ячейки для отображения пользователю."""
        if isinstance(value, str):
            return value.strip()
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

    @staticmethod
    def _sanitize_filename(name: str) -> str:
        """Удаление запрещенных символов из имени файла."""
        sanitized = re.sub(r"[<>:\"/\\\\|?*]", "_", name)
        return sanitized.strip() or "document.xlsx"

