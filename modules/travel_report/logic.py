from __future__ import annotations

"""Бизнес-логика отчета по командировке.

Содержит модели данных и экспорт отчета в Excel.
"""

from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill, NamedStyle


@dataclass
class TravelExpenseItem:
    """Статья расходов по командировке."""

    category: str
    amount: float
    document_date: Optional[date]
    document_path: Optional[str]
    cheque_number: Optional[str] = None


@dataclass
class TravelReportData:
    """Данные отчета по командировке."""

    trip_start: date
    trip_end: date
    city: str
    advance_amount: float
    items: List[TravelExpenseItem]
    companies: List[str] = field(default_factory=list)
    rewritten_note: Optional[str] = None

    def get_total_expenses(self) -> float:
        """Подсчет общей суммы расходов."""
        return sum(
            item.amount
            for item in self.items
            if item.category.lower() != "такси"
        )

    def get_balance(self) -> float:
        """Расчет остатка аванса (аванс - расходы)."""
        return self.advance_amount - self.get_total_expenses()


class TravelReportExcelExporter:
    """Экспорт отчета по командировке в Excel."""

    def __init__(self, output_directory: Path) -> None:
        """
        Инициализация экспортера.

        Args:
            output_directory: Каталог, в который будет сохранен Excel-файл.
        """
        self.output_directory = output_directory

    def export(self, report: TravelReportData, filename: str) -> Path:
        """
        Экспортирует отчет в Excel-файл.

        Args:
            report: Данные отчета.
            filename: Имя файла (без пути).

        Returns:
            Путь к созданному файлу.
        """
        self.output_directory.mkdir(parents=True, exist_ok=True)
        output_path = self.output_directory / filename

        wb = Workbook()
        ws = wb.active
        ws.title = "Отчет по командировке"

        start_row = self._write_header(ws, report)
        self._write_table_headers(ws, start_row)
        current_row = self._write_expense_items(ws, report.items, start_row + 1)
        self._write_totals(ws, report, current_row)
        self._write_note(ws, report.rewritten_note, current_row)
        self._set_column_widths(ws)

        wb.save(output_path)
        return output_path
    
    def _write_header(self, ws, report: TravelReportData) -> int:
        """Запись шапки отчета"""
        ws["A1"] = "Отчет по командировке"
        ws["A1"].font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")
        ws["A1"].alignment = Alignment(horizontal="center")

        ws["A3"] = "Город командировки:"
        ws["B3"] = report.city
        ws["A4"] = "Период:"
        ws["B4"] = f"{report.trip_start.strftime('%d.%m.%Y')} - {report.trip_end.strftime('%d.%m.%Y')}"
        ws["A5"] = "Сумма аванса:"
        ws["B5"] = report.advance_amount

        if report.companies:
            ws["A6"] = "Компания(и):"
            ws["B6"] = ", ".join(report.companies)
            return 8
        return 7
    
    def _write_table_headers(self, ws, start_row: int) -> None:
        """Запись заголовков таблицы"""
        header_fill = PatternFill("solid", fgColor="BDD7EE")
        thin_border = self._get_thin_border()

        headers = [
            "№", "Статья расходов", "Фактические расходы",
            "Номер чека", "Дата по чеку", "Файл документа",
        ]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = header_fill
            cell.border = thin_border
    
    def _write_expense_items(self, ws, items: List[TravelExpenseItem], start_row: int) -> int:
        """Запись строк расходов"""
        thin_border = self._get_thin_border()
        current_row = start_row
        
        for index, item in enumerate(items, start=1):
            row_cells = [
                ws.cell(row=current_row, column=1, value=index),
                ws.cell(row=current_row, column=2, value=item.category),
                ws.cell(row=current_row, column=3, value=item.amount),
                ws.cell(row=current_row, column=4, value=item.cheque_number),
            ]
            if item.document_date:
                row_cells.append(
                    ws.cell(row=current_row, column=5, value=item.document_date.strftime("%d.%m.%Y"))
                )
            if item.document_path:
                row_cells.append(
                    ws.cell(row=current_row, column=6, value=Path(item.document_path).name)
                )

            for cell in row_cells:
                cell.border = thin_border
                if cell.column == 3:
                    cell.number_format = "0.00"
                elif cell.column == 5:
                    cell.alignment = Alignment(horizontal="center")

            current_row += 1
        
        return current_row
    
    def _write_totals(self, ws, report: TravelReportData, current_row: int) -> None:
        """Запись итоговых сумм"""
        total_expenses = report.get_total_expenses()
        balance = report.get_balance()

        total_label_cell = ws.cell(row=current_row + 1, column=2, value="Итого расходов:")
        total_value_cell = ws.cell(row=current_row + 1, column=3, value=total_expenses)
        total_label_cell.font = Font(bold=True)
        total_value_cell.font = Font(bold=True)
        total_value_cell.number_format = "0.00"

        balance_label = self._format_balance_label(balance)
        balance_label_cell = ws.cell(row=current_row + 3, column=2, value="Итог (аванс - расходы):")
        balance_value_cell = ws.cell(row=current_row + 3, column=3, value=balance_label)
        balance_label_cell.font = Font(bold=True)
        balance_value_cell.font = Font(bold=True)
    
    def _write_note(self, ws, note: Optional[str], current_row: int) -> None:
        """Запись комментария"""
        if not note:
            return
        
        ws.cell(row=current_row + 5, column=1, value="Комментарий:")
        ws.merge_cells(
            start_row=current_row + 5,
            start_column=1,
            end_row=current_row + 7,
            end_column=5,
        )
        note_cell = ws.cell(row=current_row + 5, column=1)
        note_cell.value = note
        note_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    def _set_column_widths(self, ws) -> None:
        """Установка ширины колонок"""
        widths = {"A": 5, "B": 35, "C": 18, "D": 18, "E": 18, "F": 30}
        for col, width in widths.items():
            ws.column_dimensions[col].width = width
    
    @staticmethod
    def _get_thin_border() -> Border:
        """Создание тонкой рамки"""
        return Border(
            left=Side(style="thin", color="000000"),
            right=Side(style="thin", color="000000"),
            top=Side(style="thin", color="000000"),
            bottom=Side(style="thin", color="000000"),
        )

    @staticmethod
    def _format_balance_label(balance: float) -> str:
        """
        Форматирование итогового остатка аванса.

        Args:
            balance: Остаток (аванс - расходы).

        Returns:
            Строка в формате +X,XX / -X,XX / 0,00.
        """
        rounded = round(balance, 2)
        if rounded > 0:
            return f"+{rounded:.2f}"
        if rounded < 0:
            return f"-{abs(rounded):.2f}"
        return "0.00"


